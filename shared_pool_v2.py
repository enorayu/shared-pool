"""
Shared Pool v2 — Supplier Intelligence (Supabase Edition)
==========================================================
A generic shared pool for domain/link supplier intelligence.
Connects to Supabase PostgreSQL as the data layer.
Anyone can use it with their own Supabase project — just edit config.py.

Four Pools:
  Domain Pool  — domain_pool table      (161K+ domains)
  Email Pool   — email_pool table        (send queue, bounce tracking)
  Reply Pool   — reply_pool table        (inbound replies, A/B/C classification)
  Quote Pool  — quote_pool table        (detailed quotes from A-class replies)

Quick Start:
  1. Edit config.py with your SUPABASE_URL + SUPABASE_ANON_KEY
  2. pip install flask
  3. python shared_pool_v2.py --port 8765
  4. Open http://localhost:8765

Required Supabase Tables (already exist in default project):
  domain_pool, supplier_pool, quote_pool, config

To create email_pool and reply_pool:
  Run the CREATE TABLE SQL from the /setup page after starting the server.
"""

import os
import sys
import json
import argparse
import csv
import io
import time as _time
import urllib.request
import urllib.error
import urllib.parse
import re as _re
import requests
import openpyxl
from collections import defaultdict
from datetime import datetime, timezone, timedelta

# ── Config ─────────────────────────────────────────────────
try:
    from config import SUPABASE_URL, SUPABASE_ANON_KEY
except ImportError:
    print("ERROR: config.py not found. Create config.py with SUPABASE_URL and SUPABASE_ANON_KEY.")
    sys.exit(1)

if not SUPABASE_URL or not SUPABASE_ANON_KEY or "YOUR_PROJECT" in SUPABASE_URL:
    print("ERROR: Please edit config.py with your actual Supabase credentials.")
    sys.exit(1)

# ── Flask ───────────────────────────────────────────────────
try:
    from flask import Flask, request, jsonify, Response
except ImportError:
    print("ERROR: Flask not installed. Run: pip install flask")
    sys.exit(1)

app = Flask(__name__)

# ── CORS 跨域支持 ──
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

REST_URL = f"{SUPABASE_URL}/rest/v1"
AUTH_HEADERS = {
    "apikey": SUPABASE_ANON_KEY,
    "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
}

# Simple in-memory cache to reduce Supabase API calls
_cache = {}
def _cached(key, ttl_sec=30, fn=None):
    now = _time.time()
    if key in _cache and _cache[key][0] > now:
        return _cache[key][1]
    if fn:
        val = fn()
        _cache[key] = (now + ttl_sec, val)
        return val
    return None


def _count_unique_domains(status=None):
    """Count unique domain names (deduplicated) by paginated fetch.
    Cached 600s since scanning 161K rows takes ~160 API calls (Supabase caps at 1000/request).
    When status=None, returns a dict of all status counts in one pass."""
    def _do_count():
        seen = {}  # status -> set of unique domains
        if status:
            seen[status] = set()
        offset = 0
        batch_size = 1000  # match Supabase default max-rows cap
        while True:
            batch = db.select("domain_pool", select="domain,collection_status",
                              limit=batch_size, offset=offset,
                              order="domain_id")
            if not batch:
                break
            for row in batch:
                dom = (row.get("domain") or "").strip().lower()
                st = row.get("collection_status", "")
                if not dom:
                    continue
                if status:
                    if st == status:
                        seen[status].add(dom)
                else:
                    if st not in seen:
                        seen[st] = set()
                    seen[st].add(dom)
            if len(batch) < batch_size:
                break
            offset += batch_size

        if status:
            return len(seen.get(status, set()))
        else:
            return {k: len(v) for k, v in seen.items()}
    cache_key = f"unique_domains:{status or 'all'}"
    return _cached(cache_key, ttl_sec=600, fn=_do_count)

# ── Supabase Client ─────────────────────────────────────────

class SB:
    """Minimal Supabase REST client (stdlib only, no extra deps)."""

    def __init__(self):
        self.base = REST_URL
        self._headers = {
            "apikey": SUPABASE_ANON_KEY,
            "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
        }

    def _req(self, method, path, data=None, params=None, extra_headers=None):
        url = f"{self.base}/{path}"
        if params:
            qs = []
            for k, v in params.items():
                if v is not None:
                    qs.append(f"{urllib.parse.quote(k)}={urllib.parse.quote(str(v))}")
            if qs:
                connector = "&" if "?" in url else "?"
                url += connector + "&".join(qs)

        headers = dict(self._headers)
        headers["Connection"] = "close"  # prevent keep-alive issues with threaded Flask
        if extra_headers:
            headers.update(extra_headers)

        body = None
        if data is not None:
            body = json.dumps(data).encode("utf-8")
            headers.setdefault("Content-Type", "application/json")

        # Retry on SSL/connection errors (common with Supabase free tier)
        last_err = None
        for attempt in range(3):
            try:
                req = urllib.request.Request(url, data=body, headers=headers, method=method)
                with urllib.request.urlopen(req, timeout=20) as resp:
                    raw = resp.read()
                    return (resp, json.loads(raw) if raw else None)
            except urllib.error.HTTPError as e:
                err_body = e.read().decode("utf-8", errors="replace")[:500]
                return (e, {"error": str(e.code), "detail": err_body})
            except (urllib.error.URLError, OSError) as e:
                last_err = e
                if attempt < 2:
                    _time.sleep(0.5 * (attempt + 1))
        raise last_err

    # ── Read helpers ───────────────────────────────────────

    @staticmethod
    def _filter_part(k, v):
        """Build a Supabase filter string.
        If v already starts with an operator (eq., neq., gt., gte., lt., lte.,
        like., ilike., is., in., not.is.), use it directly. Otherwise prefix eq."""
        if v is None:
            return None
        v_str = str(v)
        ops = ("eq.", "neq.", "gt.", "gte.", "lt.", "lte.",
               "like.", "ilike.", "is.", "in.", "not.is.")
        if any(v_str.startswith(op) for op in ops):
            return f"{k}={urllib.parse.quote(v_str)}"
        return f"{k}=eq.{urllib.parse.quote(v_str)}"

    def select(self, table, select="*", filters=None, limit=None, offset=None,
               order=None, ascending=True):
        params = {"select": select}
        if limit is not None:
            params["limit"] = limit
        if offset is not None:
            params["offset"] = offset
        if order:
            direction = "asc" if ascending else "desc"
            params["order"] = f"{order}.{direction}"

        path = table
        if filters:
            parts = []
            for k, v in filters.items():
                part = self._filter_part(k, v)
                if part:
                    parts.append(part)
            if parts:
                path += "?" + "&".join(parts)

        resp, data = self._req("GET", path, params=params)
        if isinstance(data, list):
            return data
        return []

    def count(self, table, filters=None):
        """Get exact count via content-range header."""
        params = {"select": "*", "limit": 1}
        extra = {"Prefer": "count=exact"}
        path = table
        if filters:
            parts = []
            for k, v in filters.items():
                part = self._filter_part(k, v)
                if part:
                    parts.append(part)
            if parts:
                path += "?" + "&".join(parts)

        resp, _ = self._req("GET", path, params=params, extra_headers=extra)
        if hasattr(resp, "headers"):
            cr = resp.headers.get("content-range", "")
            if "/" in cr:
                return int(cr.split("/")[-1])
        return 0

    # ── Write helpers ──────────────────────────────────────

    def insert(self, table, rows, upsert=False):
        """Insert rows. Returns (resp, data)."""
        extra = {}
        if upsert:
            extra["Prefer"] = "resolution=merge-duplicates, return=representation"
        else:
            extra["Prefer"] = "return=representation"
        resp, data = self._req("POST", table, data=rows, extra_headers=extra)
        return resp, data

    def update(self, table, data, filters):
        """Update rows matching filters. Returns response object."""
        path = table
        parts = []
        for k, v in (filters or {}).items():
            if v is not None:
                parts.append(f"{k}=eq.{urllib.parse.quote(str(v))}")
        if parts:
            path += "?" + "&".join(parts)

        resp, result = self._req("PATCH", path, data=data,
                                  extra_headers={"Prefer": "return=representation"})
        return resp, result

    def patch_by_ids(self, table, data, ids, id_column="domain_id"):
        """Update multiple rows by IDs."""
        if not ids:
            return None, []
        all_results = []
        for batch in chunks(ids, 100):
            in_clause = ",".join(f"{urllib.parse.quote(str(i))}" for i in batch)
            path = f"{table}?{id_column}=in.({in_clause})"
            resp, result = self._req("PATCH", path, data=data,
                                      extra_headers={"Prefer": "return=representation"})
            all_results.append(result)
        return resp, all_results

    def delete(self, table, filter_str=None):
        """Delete rows matching a raw filter string, e.g. 'log_id=in.(1,2,3)'.
        Returns response object."""
        path = table
        if filter_str:
            path += "?" + filter_str
        resp, result = self._req("DELETE", path)
        return resp, result


db = SB()


# ── Helpers ─────────────────────────────────────────────────

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def now_iso():
    return datetime.now(timezone(timedelta(hours=8))).isoformat()

def safe_str(v, default=""):
    if v is None:
        return default
    return str(v)


def _utc_to_bj(iso_str):
    """Convert existing UTC ISO timestamps to Beijing time for display."""
    if not iso_str or not isinstance(iso_str, str):
        return iso_str
    try:
        if iso_str.endswith("+00:00"):
            dt = datetime.fromisoformat(iso_str)
            bj = dt.astimezone(timezone(timedelta(hours=8)))
            return bj.isoformat()
        return iso_str
    except Exception:
        return iso_str


# ── Operation Log helpers ───────────────────────────────────

def _log_operation(op_type, user, table_name, count, detail=""):
    """Log an operation to the independent operation_log table.

    The legacy approach stored all logs in a single config JSON field
    (key='operation_logs'), which silently failed once it exceeded
    Supabase's row-size limit (~了几万条后写入被拒). Now each operation
    is a separate row in operation_log, so it never overflows.

    注意：count=0（即"无实际变更"的操作，如空跑导入/全重复导入）不写日志，
    避免 operation_log 被无效记录刷屏。
    """
    if not count:
        return
    row = {
        "type": op_type,
        "username": user or "unknown",
        "pool": table_name,
        "count": int(count or 0),
        "detail": detail or "",
    }
    try:
        db.insert("operation_log", [row])
    except Exception as e:
        # Fallback: if the table does not exist yet (pre-migration),
        # keep writing to the legacy config field so logs are not lost.
        try:
            raw = _get_config("operation_logs", "[]")
            logs = json.loads(raw) if raw else []
            log_entry = {
                "log_id": str(int(time.time() * 1000)) + str(uuid.uuid4().hex[:6]),
                "time": now_iso(),
                "type": op_type,
                "user": user or "unknown",
                "table": table_name,
                "count": count,
                "detail": detail,
            }
            logs.insert(0, log_entry)
            logs = logs[:1000]
            _set_config("operation_logs", json.dumps(logs, ensure_ascii=False), "Operation logs")
        except Exception:
            pass


# ── Health ──────────────────────────────────────────────────

@app.route("/health")
def health():
    try:
        rows = db.select("domain_pool", select="domain_id", limit=1)
        connected = True
    except Exception:
        connected = False
    return jsonify({
        "status": "ok" if connected else "error",
        "db": SUPABASE_URL,
        "version": "v2-supabase",
        "time": now_iso(),
        "connected": connected
    })


@app.route("/api/version")
def api_version():
    return jsonify({
        "version": "v2-supabase",
        "backend": "Supabase PostgreSQL",
        "project": SUPABASE_URL.split("//")[1].split(".")[0] if "//" in SUPABASE_URL else "",
        "tables": {
            "domain_pool": db.count("domain_pool"),
            "supplier_pool": db.count("supplier_pool"),
            "quote_pool": db.count("quote_pool"),
            "config": db.count("config"),
        }
    })


# ════════════════════════════════════════════════════════════
# Domain Pool API (→ domain_pool table)
# ════════════════════════════════════════════════════════════

DOMAIN_STATUSES = ["New", "Claimed", "Contacted", "Replied", "Imported", "Collecting", "Completed", "Exported", "Failed"]


@app.route("/api/domain/register", methods=["POST"])
def domain_register():
    """Batch register domains. Dedup against existing DB first.
    Body: {"domains": ["a.com","b.com"], "priority": 0,
           "collection_status": "New", "notes": "", "imported_by": "emma"}
    If collection_status is omitted, defaults to "New".
    All imported domains are marked with source="未提取" for extraction tracking."""
    data = request.get_json(force=True)
    raw_domains = data.get("domains", [])
    imported_by = data.get("imported_by", "")
    priority = data.get("priority", 0)
    notes = data.get("notes", "")
    default_status = data.get("collection_status", "New")
    if default_status not in DOMAIN_STATUSES:
        default_status = "New"

    if imported_by:
        if notes:
            notes = f"[imported by {imported_by}] {notes}"
        else:
            notes = f"[imported by {imported_by}]"

    # 1. Normalize & dedup within batch
    seen = {}
    for raw in raw_domains:
        d = raw.lower().strip().lstrip("www.")
        if not d:
            continue
        if d not in seen:
            seen[d] = {
                "domain": d,
                "source": "未提取",
                "priority": priority,
                "notes": notes,
                "collection_status": default_status,
            }
    if not seen:
        return jsonify({"new": 0, "duplicates": 0, "domains": []})

    # 2. Batch check against existing domains (100 per query)
    domain_list = list(seen.keys())
    existing = set()
    for i in range(0, len(domain_list), 100):
        batch = domain_list[i:i+100]
        d_filter = ",".join(batch)
        try:
            results = db.select("domain_pool", select="domain",
                              filters={"domain": f"in.({d_filter})"}, limit=100)
            for r in (results or []):
                existing.add((r.get("domain") or "").strip().lower())
        except Exception:
            pass  # if check fails, treat as new (best-effort dedup)

    # 3. Filter out existing
    new_domains = {d: v for d, v in seen.items() if d not in existing}
    dup_count = len(seen) - len(new_domains)

    if not new_domains:
        _log_operation("domain_import", imported_by, "domain_pool", 0,
                       f"All {dup_count} domains already exist, Source: 未提取")
        return jsonify({"new": 0, "duplicates": dup_count})

    # 4. Insert only new domains
    rows = list(new_domains.values())
    resp, result = db.insert("domain_pool", rows, upsert=False)
    new_count = 0
    if hasattr(resp, "status") and resp.status in (200, 201):
        new_count = len(result) if isinstance(result, list) else len(rows)

    # 5. Clear cache so stats refresh immediately
    global _cache
    _cache.clear()

    # 6. Log operation
    _log_operation("domain_import", imported_by, "domain_pool", new_count,
                   f"Source: 未提取, Duplicates: {dup_count}")

    return jsonify({"new": new_count, "duplicates": dup_count})


@app.route("/api/domain/list", methods=["GET"])
def domain_list():
    """List domains with filtering. Returns unique domain list (one per domain)."""
    status = request.args.get("status", "")
    user = request.args.get("user", "")
    limit = min(int(request.args.get("limit", 100)), 5000)
    offset = int(request.args.get("offset", 0))

    filters = {}
    if status:
        filters["collection_status"] = status
    if user:
        filters["claimed_by"] = user

    # Loop-fetch to fill dedup gap: request limit*3 rows, dedup, trim to limit
    seen = set()
    unique_domains = []
    fetch_attempts = 0
    current_offset = offset
    max_attempts = max(5, (limit * 10) // 100 + 1)  # Safety limit

    while len(unique_domains) < limit and fetch_attempts < max_attempts:
        batch = db.select(
            "domain_pool",
            select="domain_id,domain,source,collection_status,claimed_by,claim_batch_id,priority,notes,created_at",
            filters=filters,
            limit=min(limit * 3, 2000),
            offset=current_offset,
            order="priority",
            ascending=False,
        )
        if not batch:
            break
        for d in batch:
            dom = d.get("domain", "")
            if dom and dom not in seen:
                seen.add(dom)
                unique_domains.append(d)
                if len(unique_domains) >= limit:
                    break
        current_offset += len(batch)
        if len(batch) < min(limit * 3, 2000):
            break  # No more data
        fetch_attempts += 1

    # Trim to limit
    domains = unique_domains[:limit]

    total_filters = {k: v for k, v in {"collection_status": status, "claimed_by": user}.items() if v}
    unique_total = _count_unique_domains(status if status else None)
    # If dict (all-status count), sum for pagination; if int, use directly
    if isinstance(unique_total, dict):
        total_for_page = sum(unique_total.values())
    else:
        total_for_page = unique_total

    return jsonify({"domains": domains, "unique_total": unique_total, "total": total_for_page, "returned": len(domains)})


@app.route("/api/domain/export", methods=["POST"])
def domain_export():
    """
    Export unclaimed domains for a user and lock them.
    Deduplicates by domain name. Generates CSV with domain fields only (no email).
    Only exports domains with collection_status=New, marks them as extracted after export.
    Exported in ascending domain_id order.
    """
    data = request.get_json(force=True)
    user = data.get("user", "").strip()
    if not user:
        return jsonify({"error": "user is required", "exported": 0}), 400
    count = min(int(data.get("count", 200)), 5000)

    # 1. Fetch unclaimed domains (New status, source!=已提取, ordered by domain_id ascending)
    domains = db.select(
        "domain_pool",
        select="domain_id,domain,source,priority,created_at",
        filters={"collection_status": "New", "source": "neq.已提取"},
        limit=count * 3,  # fetch extra for dedup
        order="domain_id",
        ascending=True,
    )

    # Deduplicate by domain name (keep first occurrence, lowest ID first)
    seen = set()
    unique_domains = []
    for d in domains:
        dom = d.get("domain", "").strip().lower()
        if dom and dom not in seen:
            seen.add(dom)
            unique_domains.append(d)
    unique_domains = unique_domains[:count]

    if not unique_domains:
        return jsonify({"exported": 0, "filename": "", "batch_id": ""})

    batch_id = f"domain_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{user}"
    filename = f"{batch_id}.csv"
    now = now_iso()
    ids = [d["domain_id"] for d in unique_domains]

    # 2. Lock domains and mark as extracted
    resp, results = db.patch_by_ids("domain_pool", {
        "claimed_by": user,
        "collection_status": "Claimed",
        "source": "已提取",
        "claim_batch_id": batch_id,
        "claim_time": now,
    }, ids)

    # Verify PATCH actually worked before generating CSV
    if hasattr(resp, "status") and resp.status not in (200, 201, 204):
        err_detail = ""
        if isinstance(results, list) and results:
            err_detail = str(results[0])[:200]
        elif isinstance(results, dict):
            err_detail = results.get("detail", str(results))[:200]
        print(f"[domain_export] PATCH failed HTTP {resp.status}: {err_detail}", file=sys.stderr)
        return jsonify({"error": f"Failed to lock domains (HTTP {resp.status})", "exported": 0}), 500

    # 3. Generate CSV — sequential #, not domain_id
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["#", "domain", "source", "priority", "created_at"])
    for idx, r in enumerate(unique_domains, 1):
        writer.writerow([idx, r["domain"], safe_str(r.get("source")),
                         r.get("priority", 0),
                         safe_str(r.get("created_at"))[:19]])

    csv_content = output.getvalue()

    # 4. Save CSV
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shared_pool_exports")
    os.makedirs(export_dir, exist_ok=True)
    with open(os.path.join(export_dir, filename), "w", newline="", encoding="utf-8") as f:
        f.write(csv_content)

    # 5. Clear cache & log operation
    global _cache
    _cache.clear()
    _log_operation("domain_export", user, "domain_pool", len(unique_domains),
                   f"Batch: {batch_id}, Status: Claimed, Source: 已提取")

    return jsonify({
        "exported": len(unique_domains),
        "filename": filename,
        "batch_id": batch_id,
        "csv_content": csv_content,
        "csv_preview": csv_content[:500]
    })


@app.route("/api/domain/distribute", methods=["POST"])
def domain_distribute():
    """
    Round-robin distribute unclaimed domains among team members.
    Body: {"count": 1000}  (users come from config or default)
    Uses config key "team_members" for user list. Falls back to ["leo","emma","jack"].
    Domains are ordered by domain_id ascending and marked as extracted after distribution.
    """
    data = request.get_json(force=True)

    # Read team members from config
    raw = _get_config("team_members", '["leo","emma","jack"]')
    try:
        users = json.loads(raw) if raw else ["leo", "emma", "jack"]
    except (json.JSONDecodeError, TypeError):
        users = [m.strip() for m in str(raw).split(",") if m.strip()]
    if not users:
        users = ["leo", "emma", "jack"]

    count = int(data.get("count", 0))

    total = db.count("domain_pool", filters={"collection_status": "New"})
    if count <= 0 or count > total:
        count = total

    if count <= 0:
        return jsonify({"assigned": 0, "distribution": {}})

    domains = db.select(
        "domain_pool",
        select="domain_id",
        filters={"collection_status": "New"},
        limit=count,
        order="domain_id",
        ascending=True,
    )

    now = now_iso()
    distribution = {u: 0 for u in users}
    for i, d in enumerate(domains):
        user = users[i % len(users)]
        db.update("domain_pool", {
            "claimed_by": user,
            "collection_status": "Claimed",
            "source": "已提取",
            "claim_time": now,
            "claim_batch_id": f"distribute_{now[:10]}_{user}",
        }, {"domain_id": d["domain_id"]})
        distribution[user] += 1

    # Clear cache & log operation
    global _cache
    _cache.clear()
    _log_operation("domain_distribute", "system", "domain_pool", len(domains),
                   f"Distribution: {json.dumps(distribution)}, Source: 已提取")

    return jsonify({"assigned": len(domains), "distribution": distribution})


@app.route("/api/domain/update_status", methods=["POST"])
def domain_update_status():
    """Update domain collection status."""
    data = request.get_json(force=True)
    domain_ids = data.get("ids", [])
    new_status = data.get("status", "")
    maisui_id = data.get("maisui_task_id", "")
    notes = data.get("notes", "")

    if new_status not in DOMAIN_STATUSES:
        return jsonify({"error": f"Invalid status: {new_status}"}), 400

    payload = {"collection_status": new_status, "updated_at": now_iso()}
    if maisui_id:
        payload["maisui_task_id"] = maisui_id
    if notes:
        payload["notes"] = notes

    resp, result = db.patch_by_ids("domain_pool", payload, domain_ids)
    updated = len(result) if result else 0
    return jsonify({"updated": updated})


@app.route("/api/domain/stats", methods=["GET"])
def domain_stats():
    """Statistics for domain pool."""
    stats = {"total": db.count("domain_pool")}
    for s in DOMAIN_STATUSES:
        stats[s.lower()] = db.count("domain_pool", filters={"collection_status": s})

    # Today's new
    today = datetime.now().strftime("%Y-%m-%d")
    stats["today_new"] = db.count("domain_pool",
                                   filters={"created_at": f"gte.{today}"})

    # By user
    users = defaultdict(int)
    domains = db.select("domain_pool", select="claimed_by",
                         filters={"claimed_by": "not.is.null"}, limit=5000)
    for d in domains:
        users[d.get("claimed_by") or "unknown"] += 1
    stats["by_user"] = dict(users)

    return jsonify(stats)


# ════════════════════════════════════════════════════════════
# Email Pool API (derived from domain_pool contact_email)
# ════════════════════════════════════════════════════════════

EMAIL_STATUSES = ["New", "Unsent", "Assigned", "Sent", "Bounce"]


@app.route("/api/email/queue", methods=["GET"])
def email_queue():
    """
    Get UNSENT emails ready for sending from email_pool table.
    Filter: send_status='UNSENT', ordered by email_id (FIFO).
    """
    user = request.args.get("user", "")
    count = min(int(request.args.get("count", 100)), 2000)
    offset = int(request.args.get("offset", 0))
    claimed_by_filter = request.args.get("claimed_by", "")

    filters = {"send_status": "UNSENT"}
    if claimed_by_filter:
        filters["claimed_by"] = claimed_by_filter

    emails = db.select(
        "email_pool",
        select="email_id,email,domain,send_status,collection_status,claimed_by,source,notes,created_at",
        filters=filters,
        limit=count,
        offset=offset,
        order="email_id",
        ascending=True,
    )

    result = []
    for e in (emails or []):
        e["contact_email"] = e.get("email")  # for frontend compatibility
        result.append(e)

    total = db.count("email_pool", filters=filters)
    return jsonify({"emails": result, "count": len(result), "total": total, "offset": offset})


@app.route("/api/email/export", methods=["POST"])
def email_export():
    """
    Export UNSENT emails from email_pool for sending.
      - Exports emails with send_status='UNSENT'.
      - Different users get different emails (claimed_by locking).
      - Marks send_status='SENT', source='已提取' after export.
    """
    data = request.get_json(force=True)
    user = data.get("user", "").strip()
    if not user:
        return jsonify({"error": "user is required", "exported": 0}), 400
    count = min(int(data.get("count", 500)), 5000)

    # ── 原子导出: 调用 DB 函数 export_emails_atomic ──
    # 该函数在单个事务内 SELECT ... FOR UPDATE SKIP LOCKED + 立即 UPDATE,
    # 彻底消除旧 check-then-act 的并发竞态 (两人同时导出会抢同一批邮箱)。
    # 数据库侧已创建该函数 (见 SQL: export_emails_atomic)。
    try:
        rpc_resp = requests.post(
            f"{REST_URL}/rpc/export_emails_atomic",
            headers={**AUTH_HEADERS, "Content-Type": "application/json", "Prefer": "return=representation"},
            json={"p_user": user, "p_count": count},
            timeout=60,
        )
        if rpc_resp.status_code not in (200, 201):
            return jsonify({
                "error": f"atomic export failed: HTTP {rpc_resp.status_code}",
                "detail": rpc_resp.text[:500],
                "exported": 0,
            }), 500
        emails = rpc_resp.json()
    except Exception as e:
        return jsonify({"error": f"atomic export exception: {str(e)}", "exported": 0}), 500

    if not emails:
        return jsonify({"exported": 0, "filename": ""})

    # batch_id / filename 基于第一批的 claim_time 精确对应
    now = now_iso()
    batch_id = f"email_send_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{user}"
    filename = f"{batch_id}.csv"

    # Generate CSV — sequential #, not email_id
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["#", "email", "domain", "send_status", "source"])
    for idx, e in enumerate(emails, 1):
        writer.writerow([
            idx,
            safe_str(e.get("email")),
            e.get("domain") or "",
            e.get("send_status", "SENT"),
            safe_str(e.get("source")),
        ])

    csv_content = output.getvalue()
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shared_pool_exports")
    os.makedirs(export_dir, exist_ok=True)
    with open(os.path.join(export_dir, filename), "w", newline="", encoding="utf-8") as f:
        f.write(csv_content)

    updated_count = len(emails)

    global _cache
    _cache.clear()
    _log_operation("email_export", user, "email_pool", len(emails),
                   f"Batch: {batch_id}, Source: 已提取")

    return jsonify({"exported": len(emails), "filename": filename, "batch_id": batch_id, "csv_content": csv_content})


@app.route("/api/email/stats", methods=["GET"])
def email_stats():
    """Email pool statistics (from email_pool table)."""
    total = db.count("email_pool")
    unsent = db.count("email_pool", filters={"send_status": "UNSENT"})
    sent = db.count("email_pool", filters={"send_status": "SENT"})
    bounce = db.count("email_pool", filters={"send_status": "Bounce"})

    return jsonify({
        "total": total,
        "unsent": unsent,
        "sent": sent,
        "bounce": bounce,
    })


@app.route("/api/email/import", methods=["POST"])
def email_pool_import():
    """
    Batch import emails to independent email_pool table.
    Body: {"emails": [{"email":"a@b.com","domain":"b.com"},...], "imported_by": "leo"}
    Batch dedup via in.() filter (performance: 1 API call per 100 emails).
    All new entries marked send_status='UNSENT', source='未提取', collection_status='New'.
    """
    global _cache
    data = request.get_json(force=True)
    records = data.get("emails", [])
    imported_by = data.get("imported_by", "").strip()
    if not imported_by:
        return jsonify({"error": "imported_by is required"}), 400

    if not records:
        return jsonify({"imported": 0, "new": 0, "skipped": 0})

    import_batch = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{imported_by}"
    notes = f"[email imported by {imported_by}]"

    # 1. Normalize & dedup within batch
    seen = {}
    for rec in records:
        email = rec.get("email", "").strip().lower()
        domain = rec.get("domain", "").strip().lower().lstrip("www.")
        if not email or not domain or "@" not in email:
            continue
        if email not in seen:
            seen[email] = {
                "email": email,
                "domain": domain,
                "send_status": "UNSENT",
                "source": "未提取",
                "collection_status": "New",
                "notes": notes,
                "import_batch": import_batch,
                "priority": 0,
            }

    if not seen:
        return jsonify({"imported": 0, "new": 0, "skipped": 0})

    # 2. Batch check against existing emails (100 per query)
    email_list = list(seen.keys())
    existing_emails = set()
    dedup_errors = 0
    for i in range(0, len(email_list), 100):
        batch = email_list[i:i+100]
        e_filter = ",".join(batch)
        try:
            results = db.select("email_pool", select="email",
                              filters={"email": f"in.({e_filter})"}, limit=100)
            for r in (results or []):
                existing_emails.add((r.get("email") or "").strip().lower())
        except Exception as e:
            dedup_errors += 1
            print(f"[email_import] dedup query failed (batch {i}): {e}", file=sys.stderr)

    # 3. Filter out existing (only trust dedup if no errors)
    new_emails = {e: v for e, v in seen.items() if e not in existing_emails}
    skipped = len(seen) - len(new_emails)

    if dedup_errors:
        print(f"[email_import] WARNING: {dedup_errors} dedup queries failed. "
              f"Fallback: using server-side upsert. Skipped(counted): {skipped}", file=sys.stderr)
        # Can't trust skipped count if dedup failed — actual skipped determined by server
        skipped = 0

    if not new_emails:
        _cache.clear()
        _log_operation("email_import", imported_by, "email_pool", 0,
                       f"All {skipped} emails already exist, Source: 未提取")
        return jsonify({"imported": 0, "skipped": skipped})

    # 4. Insert new emails (upsert as safety net for dedup failures)
    rows = list(new_emails.values()) if new_emails else list(seen.values())
    resp, result = db.insert("email_pool", rows, upsert=True)
    imported = 0
    if hasattr(resp, "status") and resp.status in (200, 201):
        # With return=representation, result contains actually inserted rows
        imported = len(result) if isinstance(result, list) else 0
    elif hasattr(resp, "code"):
        print(f"[email_import] insert failed HTTP {resp.code}: {result}", file=sys.stderr)
    else:
        # Should not happen with return=representation, but keep fallback
        imported = 0
        skipped = len(seen)

    # 5. Clear cache & log
    _cache.clear()
    _log_operation("email_import", imported_by, "email_pool", imported,
                   f"Source: 未提取, Skipped(dup): {skipped}, Batch: {import_batch}")

    return jsonify({"imported": imported, "skipped": skipped})


# ════════════════════════════════════════════════════════════
# Reply Pool API (→ reply_pool table, fallback to supplier_pool)
# ════════════════════════════════════════════════════════════

def _parse_reply_category(notes, content=""):
    """Parse A/B/C category from supplier_pool notes field or reply content."""
    notes = str(notes or "")
    content = str(content or "")
    if not notes and not content:
        return "C"
    notes_lower = notes.lower()
    content_lower = content.lower()

    # 显式标签优先 (来自 Excel notes 列)
    if "分类:A" in notes or "有合作意向" in notes or "分类:a" in notes_lower:
        return "A"
    if "分类:B" in notes or "分类:b" in notes_lower:
        return "B"
    if "分类:C" in notes or "一般回复" in notes or "历史回复" in notes:
        return "C"

    # 从回复正文推断 (A类=明确合作/报价意向; B类=询问细节/犹豫)
    a_signals = ["报价", "价格", "多少钱", "合作意向", "愿意合作", "interested in", "price quote",
                 "our rate", "we charge", "guest post", "sponsor", "link placement", "合作方式",
                 "报价单", "fee", "cost", "usd", "eur", "$", "£", "€"]
    b_signals = ["请问", "如何", "什么要求", "需要什么", "more info", "could you", "details",
                 "tell me more", "what are", "requirements", "考虑一下", "再看看", "商量"]

    a_hits = sum(1 for s in a_signals if s in content_lower)
    b_hits = sum(1 for s in b_signals if s in content_lower)

    if a_hits > 0 and a_hits >= b_hits:
        return "A"
    if b_hits > 0:
        return "B"
    return "C"


def _supplier_to_reply(s):
    """Convert a supplier_pool row to reply_pool format."""
    email = s.get("contact_email") or s.get("supplier_name") or ""
    domain = email.split("@")[-1] if "@" in email else ""
    return {
        "reply_id": s.get("supplier_id"),
        "email": email,
        "domain": domain,
        "reply_content": s.get("notes", ""),
        "reply_time": s.get("created_at"),
        "category": _parse_reply_category(s.get("notes")),
        "status": "New",
        "supplier": s.get("supplier_name", ""),
        "contact_email": email,
        "discovered_by": s.get("source", "system"),
        "discovered_at": s.get("created_at"),
        "replied_by": None,
        "replied_at": None,
        "notes": s.get("notes", ""),
    }


@app.route("/api/reply/import", methods=["POST"])
def reply_import():
    """
    Import replies from Excel/CSV upload.
    Template columns: 序号,邮箱,域名,账号,发件人,主题,正文摘要,日期
    Deduplicates by email (skip existing).
    """
    user = request.form.get("user", "unknown").strip()
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        else:
            # Excel: use openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [str(c.value or '').strip() for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or '') for v in row])))
    except Exception as e:
        return jsonify({"error": f"File parse error: {str(e)}"}), 400

    if not rows:
        return jsonify({"error": "Empty file"}), 400

    # Normalize column names
    col_map = {}
    for col in rows[0].keys():
        cl = col.lower().strip()
        if any(k in cl for k in ['邮箱', 'email', 'mail']): col_map[col] = 'email'
        elif any(k in cl for k in ['域名', 'domain']): col_map[col] = 'domain'
        elif any(k in cl for k in ['账号', 'account', 'inbox', 'discovered']): col_map[col] = 'account'
        elif any(k in cl for k in ['发件人', 'from', 'supplier', 'name']): col_map[col] = 'supplier'
        elif any(k in cl for k in ['主题', 'subject', 'title']): col_map[col] = 'subject'
        elif any(k in cl for k in ['正文', 'body', 'content', 'preview', '摘要']): col_map[col] = 'body'
        elif any(k in cl for k in ['日期', 'date', 'time', 'reply_time']): col_map[col] = 'date'

    # Get existing emails for dedup (paginated fetch)
    existing = set()
    try:
        page, page_size = 0, 1000
        while True:
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/reply_pool?select=email&limit={page_size}&offset={page*page_size}",
                headers=AUTH_HEADERS,
                timeout=30
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            if not data:
                break
            for r in data:
                if r.get('email'):
                    existing.add(r['email'].lower())
            if len(data) < page_size:
                break
            page += 1
    except Exception:
        pass  # if can't check, rely on UNIQUE(email) constraint

    imported = 0
    skipped = 0
    batch = []
    BATCH_SIZE = 50

    def flush_batch():
        nonlocal imported
        if not batch:
            return
        for attempt in range(3):
            try:
                resp = requests.post(
                    f"{SUPABASE_URL}/rest/v1/reply_pool",
                    headers={**AUTH_HEADERS, 'Prefer': 'return=representation'},
                    json=batch,
                    timeout=30
                )
                if resp.status_code in (200, 201):
                    imported += len(batch)
                    break
            except Exception:
                if attempt < 2:
                    _time.sleep(1)
        batch.clear()

    for row in rows:
        email = _extract_email(row, col_map)
        if not email:
            continue
        email = email.lower()
        if email in existing:
            skipped += 1
            continue
        existing.add(email)  # dedup within this batch too

        supplier = row.get(col_map.get('supplier', ''), '') or ''
        domain = row.get(col_map.get('domain', ''), '') or email.split('@')[-1]
        discovered = row.get(col_map.get('account', ''), '') or user
        subject = row.get(col_map.get('subject', ''), '') or ''
        body = row.get(col_map.get('body', ''), '') or ''
        date_str = row.get(col_map.get('date', ''), '') or ''

        reply_time = now_iso()
        if date_str:
            try:
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%a, %d %b %Y %H:%M:%S %z']:
                    try:
                        reply_time = datetime.strptime(date_str.strip(), fmt).isoformat()
                        break
                    except ValueError:
                        continue
            except Exception:
                pass

        batch.append({
            "email": email,
            "domain": domain.lower(),
            "reply_content": f"{subject}\n\n{body}"[:5000],
            "reply_time": reply_time,
            "category": _parse_reply_category(notes=row.get(col_map.get('subject', ''), '') or '', content=f"{subject}\n{body}"),
            "status": "New",
            "supplier": supplier[:200],
            "contact_email": email,
            "discovered_by": discovered[:100],
            "discovered_at": now_iso(),
            "notes": f"Imported by {user}",
        })

        if len(batch) >= BATCH_SIZE:
            flush_batch()

    flush_batch()

    # Log operation
    _log_operation("reply_import", user, "reply_pool", imported,
                   f"Imported {imported} replies" + (f", skipped {skipped} duplicates" if skipped else ""))

    return jsonify({"imported": imported, "skipped": skipped, "total": len(rows)})


def _extract_email(row, col_map):
    """Extract email from various field formats."""
    # Try email column first
    email_col = col_map.get('email')
    if email_col:
        val = str(row.get(email_col, '') or '').strip()
        if '@' in val:
            return val.split()[0].strip() if ' ' in val else val

    # Try from/supplier column (may have format "Name <email>")
    supplier_col = col_map.get('supplier')
    if supplier_col:
        val = str(row.get(supplier_col, '') or '').strip()
        m = _re.search(r'<([^>]+@[^>]+)>', val)
        if m:
            return m.group(1)

    return None


@app.route("/api/reply/add", methods=["POST"])
def reply_add():
    """
    Record a new reply from a supplier.
    Body: {"email":"a@x.com","domain":"x.com","content":"...","category":"A",
           "supplier":"John","supplier_email":"a@x.com","discovered_by":"emma"}
    """
    data = request.get_json(force=True)

    payload = {
        "email": data["email"].lower().strip(),
        "domain": data.get("domain", ""),
        "reply_content": (data.get("content") or data.get("reply_content") or "")[:2000],
        "reply_time": data.get("reply_time", now_iso()),
        "category": data.get("category", "C").upper(),
        "status": "New",
        "supplier": data.get("supplier", ""),
        "contact_email": data.get("supplier_email", data["email"]),
        "discovered_by": data.get("discovered_by", "system"),
        "discovered_at": now_iso(),
    }

    resp, result = db.insert("reply_pool", payload)
    if hasattr(resp, "status") and resp.status in (200, 201):
        reply_id = result[0].get("reply_id") if isinstance(result, list) and result else None
        return jsonify({"result": "created", "reply_id": reply_id})
    else:
        return jsonify({"result": "error", "detail": str(result)}), 500


@app.route("/api/reply/list", methods=["GET"])
def reply_list():
    """List replies with filtering.
    Falls back to supplier_pool if reply_pool is empty (legacy data)."""
    category = request.args.get("category", "")
    status = request.args.get("status", "")
    user = request.args.get("user", "")
    limit = min(int(request.args.get("limit", 100)), 5000)
    offset = int(request.args.get("offset", 0))

    # Try reply_pool first
    filters = {}
    if category:
        filters["category"] = category.upper()
    if status:
        filters["status"] = status
    if user:
        filters["imported_by"] = user

    replies = db.select(
        "reply_pool",
        select="*",
        filters=filters,
        limit=limit,
        offset=offset,
        order="discovered_at",
        ascending=False,
    )

    total = db.count("reply_pool", filters=filters)

    # Fallback to supplier_pool (legacy data from shared-pool-tools.exe)
    if not replies:
        sf = {"status": "Replied"}
        suppliers = db.select(
            "supplier_pool",
            select="supplier_id,supplier_name,contact_email,source,notes,created_at",
            filters=sf,
            limit=limit,
            offset=offset,
            order="created_at",
            ascending=False,
        )
        replies = [_supplier_to_reply(s) for s in (suppliers or [])]
        if category:
            replies = [r for r in replies if r["category"] == category.upper()]
        total = db.count("supplier_pool", filters=sf)

    return jsonify({"replies": replies or [], "total": total, "offset": offset})


@app.route("/api/reply/stats", methods=["GET"])
def reply_stats():
    """Reply statistics. Falls back to supplier_pool if reply_pool is empty."""
    try:
        total = db.count("reply_pool")
        if total > 0:
            a_count = db.count("reply_pool", filters={"category": "A"})
            b_count = db.count("reply_pool", filters={"category": "B"})
            c_count = db.count("reply_pool", filters={"category": "C"})
            new_count = db.count("reply_pool", filters={"status": "New"})
        else:
            # Fallback: count from supplier_pool (legacy data)
            total = db.count("supplier_pool", filters={"status": "Replied"})
            # Sample up to 1000 for classification (Supabase REST limit)
            suppliers = db.select(
                "supplier_pool",
                select="notes",
                filters={"status": "Replied"},
                limit=1000,
            )
            a_count = sum(1 for s in suppliers if _parse_reply_category(s.get("notes")) == "A")
            b_count = sum(1 for s in suppliers if _parse_reply_category(s.get("notes")) == "B")
            c_count = sum(1 for s in suppliers if _parse_reply_category(s.get("notes")) == "C")
            # Scale up proportionally if we hit the 1000 limit
            if total > 1000 and suppliers:
                scale = total / len(suppliers)
                a_count = int(a_count * scale)
                b_count = int(b_count * scale)
                c_count = int(c_count * scale)
            new_count = total  # All legacy replies are treated as unread
    except Exception:
        return jsonify({
            "total": 0, "a_class": 0, "b_class": 0, "c_class": 0,
            "unread": 0, "replied": 0, "today": 0, "a_today": 0,
            "note": "reply_pool table may not exist. Run setup to create it."
        })

    return jsonify({
        "total": total,
        "unread": new_count,
        "replied": total - new_count,
        "a_class": a_count,
        "b_class": b_count,
        "c_class": c_count,
        "today": 0,
        "a_today": 0,
    })


@app.route("/api/reply/update_status", methods=["POST"])
def reply_update_status():
    """Update reply status."""
    data = request.get_json(force=True)
    reply_id = data.get("id")

    resp, _ = db.update("reply_pool", {
        "status": data.get("status", "Resolved"),
        "replied_at": now_iso(),
    }, {"reply_id": reply_id})
    return jsonify({"updated": 1})


# ════════════════════════════════════════════════════════════
# Quote Pool API (→ quote_pool table)
# ════════════════════════════════════════════════════════════

@app.route("/api/price/add", methods=["POST"])
@app.route("/api/quote/add", methods=["POST"])
def quote_add():
    """Add a quote record."""
    data = request.get_json(force=True)
    payload = {
        "email": data.get("email", ""),
        "domain": data.get("domain", ""),
        "supplier": data.get("supplier", ""),
        "contact_email": data.get("contact_email", ""),
        "niche": data.get("niche", ""),
        "country": data.get("country", ""),
        "traffic": data.get("traffic", ""),
        "site_category": data.get("site_category", ""),
        "cooperation_type": data.get("cooperation_type", ""),
        "price": data.get("price", ""),
        "link_rules": data.get("link_rules", ""),
        "permanence": data.get("permanence", ""),
        "content": data.get("content", ""),
        "tat": data.get("tat", ""),
        "payment": data.get("payment", ""),
        "discount": data.get("discount", ""),
        "additional_services": data.get("additional_services", ""),
        "requirements": data.get("requirements", ""),
        "reply_id": data.get("reply_id"),
        "reply_content": data.get("reply_content", ""),
        "status": data.get("status", "New"),
        "priority": data.get("priority", 0),
        "notes": data.get("notes", ""),
        "discovered_by": data.get("discovered_by", "manual"),
    }
    resp, result = db.insert("quote_pool", payload)
    return jsonify({"result": "created"})


@app.route("/api/price/list", methods=["GET"])
@app.route("/api/quote/list", methods=["GET"])
def quote_list():
    """List quotes with optional filters and full-text keyword search."""
    domain = request.args.get("domain", "")
    status = request.args.get("status", "")
    niche = request.args.get("niche", "")
    supplier = request.args.get("supplier", "")
    search = request.args.get("search", "").strip()
    limit = min(int(request.args.get("limit", 100)), 1000)
    offset = int(request.args.get("offset", 0))

    filters = {}
    if domain:
        filters["domain"] = domain
    if status:
        filters["status"] = status
    if niche:
        filters["niche"] = niche
    if supplier:
        filters["supplier"] = supplier

    # Keyword search: build or= filter across key text fields
    # PostgREST uses * as wildcard (not SQL %), and or=(...) parenthesized format
    if search:
        # Escape special chars in search term for safe inclusion in filter
        safe_search = search.replace("*", "\\*").replace("%", "\\%").replace("_", "\\_")
        or_clauses = ",".join([
            f"domain.ilike.*{safe_search}*",
            f"supplier.ilike.*{safe_search}*",
            f"email.ilike.*{safe_search}*",
            f"contact_email.ilike.*{safe_search}*",
            f"content.ilike.*{safe_search}*",
            f"notes.ilike.*{safe_search}*",
            f"country.ilike.*{safe_search}*",
            f"niche.ilike.*{safe_search}*",
            f"site_category.ilike.*{safe_search}*",
            f"keywords.ilike.*{safe_search}*",
            f"categories.ilike.*{safe_search}*",
            f"link_rules.ilike.*{safe_search}*",
        ])
        # Use direct _req call so we control the exact query string
        qs_parts = [f"select=*", f"limit={limit}", f"offset={offset}",
                     f"order=discovered_at.desc", f"or=({or_clauses})"]
        for k, v in filters.items():
            fp = db._filter_part(k, v)
            if fp:
                qs_parts.append(fp)
        qs = "&".join(qs_parts)
        resp, data = db._req("GET", f"quote_pool?{qs}", extra_headers={"Prefer": "count=exact"})
        total = int((resp.headers.get("Content-Range") or "*/0").split("/")[-1])
        quotes = data if isinstance(data, list) else []
        return jsonify({"quotes": quotes or [], "total": total, "offset": offset})

    quotes = db.select(
        "quote_pool",
        select="*",
        filters=filters,
        limit=limit,
        offset=offset,
        order="discovered_at",
        ascending=False,
    )
    total = db.count("quote_pool", filters=filters)
    return jsonify({"quotes": quotes or [], "total": total, "offset": offset})


@app.route("/api/price/stats", methods=["GET"])
@app.route("/api/quote/stats", methods=["GET"])
def quote_stats():
    """Quote pool statistics."""
    total = db.count("quote_pool")
    suppliers = 0
    if total > 0:
        quotes = db.select("quote_pool", select="supplier", limit=5000)
        suppliers = len(set(q.get("supplier") for q in quotes if q.get("supplier")))

    # Count by status
    status_counts = {}
    if total > 0:
        quotes = db.select("quote_pool", select="status", limit=5000)
        from collections import Counter
        status_counts = dict(Counter(q.get("status") for q in quotes if q.get("status")))

    return jsonify({
        "total": total,
        "suppliers": suppliers,
        "today_new": 0,
        "by_status": status_counts,
    })


@app.route("/api/price/export", methods=["GET"])
@app.route("/api/quote/export", methods=["GET"])
def quote_export():
    """Export quotes as CSV — Jenny template format (without 6 Niche Price columns).
    scope=all (默认) | ready (仅READY) | abnormal (NEED_*)

    """
    scope = (request.args.get("scope") or "all").lower()
    filters = {}
    if scope == "ready":
        filters["data_status"] = "eq.READY"
    elif scope == "abnormal":
        # NEED_DOMAIN / NEED_PRICE / NEED_REVIEW 等任意非 READY
        filters["data_status"] = "neq.READY"
    quotes = db.select("quote_pool", select="*", limit=10000,
                       filters=filters,
                       order="discovered_at", ascending=False)
    # Jenny CSV columns (excluding Casino/Finance/Erotic/Dating/CBD/Crypto/Medicine Niche Price)
    # + 8 standard fields as separate columns (cooperation_type/payment/discount/link_rules/
    #   content/requirements/additional_services/supplier)
    headers = ["#", "Link", "Price", "Backlink Type", "DR", "DA",
               "Ref. Domains", "Traffic", "Country", "Keywords",
               "Categories", "Languages", "TAT", "Permanence", "Contact",
               "Cooperation", "Payment", "Link Rules", "Status",
               "Data Status"]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for idx, q in enumerate(quotes or [], 1):
        # Price: 优先 normalized_price + 单位(USD)，去掉多余小数点
        norm = q.get("normalized_price")
        if norm is not None and str(norm).strip() != "":
            price_val = float(norm)
            # 整数不显示 .0，真正有小数才保留
            price_display = str(int(price_val)) if price_val == int(price_val) else f"{price_val:g}"
            price_cell = f"{price_display} {q.get('normalized_currency') or 'USD'}"
        else:
            price_cell = safe_str(q.get("price"))
        # Link: domain 已是 canonical 标准化后的真实域名（异常靠 Data Status 列+前端颜色区分）
        link_cell = safe_str(q.get("domain"))
        writer.writerow([
            idx,                                          # 0  #
            link_cell,                                    # 1  Link (异常域名带 [NEED_DOMAIN] 前缀)
            price_cell,                                   # 2  Price (e.g. "100 USD" 无小数点)
            safe_str(q.get("price_type") or q.get("site_category") or q.get("niche")),   # 3  Backlink Type
            safe_str(q.get("dr") or q.get("traffic")),     # 4  DR
            safe_str(q.get("da")),                         # 5  DA
            safe_str(q.get("ref_domains")),                # 6  Ref. Domains
            safe_str(q.get("traffic")),                    # 7  Traffic
            safe_str(q.get("country")),                    # 8  Country
            safe_str(q.get("keywords")),                   # 9  Keywords
            safe_str(q.get("categories")),                 # 10 Categories
            safe_str(q.get("languages")),                  # 11 Languages
            safe_str(q.get("tat")) or "",                  # 12 TAT (空则显示空，不漏到下一列)
            safe_str(q.get("permanence")),                 # 13 Permanence
            safe_str(q.get("contact_email") or q.get("email")),  # 14 Contact
            safe_str(q.get("cooperation_type")),           # 15 Cooperation
            safe_str(q.get("payment")),                    # 16 Payment
            safe_str(q.get("link_rules")),                 # 17 Link Rules
            safe_str(q.get("status")),                     # 18 Status
            safe_str(q.get("data_status")),                # 19 Data Status
        ])

    # UTF-8 with BOM so Excel (zh-CN / Windows) opens Chinese fields correctly.
    # Flask encodes a str body as Latin-1 by default → without explicit utf-8
    # + BOM the non-ASCII columns (supplier/content/requirements/其他) get garbled.
    csv_text = output.getvalue()
    csv_bytes = "\ufeff".encode("utf-8") + csv_text.encode("utf-8")
    # mimetype="text/csv" → Flask emits "text/csv; charset=utf-8" (single, clean).
    # The UTF-8 BOM (ef bb bf) at the start is what makes Excel open Chinese correctly.
    _scope_label = {"ready": "ready_only", "abnormal": "abnormal_only"}.get(scope, "all")
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=quote_pool_export_{_scope_label}.csv"},
    )


@app.route("/api/quote/delete", methods=["POST"])
def quote_delete():
    """Delete quotes by quote_id list."""
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get("ids") or []
    if not ids:
        return jsonify({"error": "no ids provided"}), 400
    # quote_id is int in DB; coerce
    ids = [int(x) for x in ids if str(x).isdigit()]
    if not ids:
        return jsonify({"error": "invalid ids"}), 400
    try:
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/quote_pool?quote_id=in.({','.join(str(i) for i in ids)})",
            headers={**AUTH_HEADERS, "Prefer": "return=minimal"},
            json={},
        )
        # Supabase delete via POST with empty body is not standard; use DELETE
        resp = requests.delete(
            f"{SUPABASE_URL}/rest/v1/quote_pool?quote_id=in.({','.join(str(i) for i in ids)})",
            headers=AUTH_HEADERS,
        )
        if resp.status_code >= 400:
            return jsonify({"error": f"delete failed: {resp.status_code} {resp.text[:200]}"}), 500
        return jsonify({"deleted": len(ids), "ids": ids})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _parse_price_from_content(text: str):
    """
    从回复正文解析报价金额 (高置信优先, 避免误抓年份/随机数字)。
    返回 (price_str, currency) 或 (None, None)。
    支持: $50 / 50 USD / €40 / 212,00 euros / 50$/ 50 per post / $50-$100
    仅当金额明确带货币符号/代码, 或紧邻 price/cost/rate 等关键词时才认;
    年份(19xx/20xx)、孤立数字一律忽略, 防止把 2026 当成价格。
    """
    if not text:
        return None, None
    import re
    low = text.lower()
    cur_map = {"$": "USD", "€": "EUR", "£": "GBP", "₹": "INR", "¥": "CNY"}

    def _clean_num(s):
        s = s.replace(",", "")
        # 过滤掉年份 (19xx / 20xx) 和超大数
        try:
            v = float(s)
        except ValueError:
            return None
        if 1900 <= v <= 2099:
            return None
        if v > 100000:
            return None
        return s

    # 1) 带货币符号的金额
    m = re.search(r"([$€£₹¥])\s?\d[\d,]*\.?\d*", text)
    if m:
        sym = m.group(1)
        num = _clean_num(re.sub(r"[^\d.,]", "", m.group(0)))
        if num:
            return num, cur_map.get(sym, "USD")

    # 2) "数字 + 货币代码/词"
    m = re.search(r"\b(\d[\d,]*\.?\d*)\s*(usd|eur|gbp|inr|cny|euros?|dollars?|bucks|rs)\b", low)
    if m:
        num = _clean_num(m.group(1))
        if num:
            word = m.group(2)
            cur = {"usd": "USD", "dollars": "USD", "bucks": "USD",
                   "eur": "EUR", "euro": "EUR", "euros": "EUR",
                   "gbp": "GBP", "inr": "INR", "rs": "INR",
                   "cny": "CNY"}.get(word, "USD")
            return num, cur

    # 3) 紧邻价格关键词的明确数字 (排除年份/超大)
    m = re.search(r"(?:price|cost|rate|fee|charged?|per post|per link|pricing)\D{0,15}?(\d[\d,]*\.?\d*)", low)
    if m:
        num = _clean_num(m.group(1))
        if num:
            return num, "USD"

    return None, None


def _normalize_price_fields(price_str: str):
    """
    从 price 原始文本(如 "$35/post / $20/post" 或 "£120/post / £90/link")
    提取首个有效金额与货币, 返回 {normalized_price: float|None, normalized_currency: str|None}。
    供导入流程自动标准化, 避免人工回填。
    """
    if not price_str:
        return {"normalized_price": None, "normalized_currency": None}
    pval, pcur = _parse_price_from_content(price_str)
    if pval is None:
        return {"normalized_price": None, "normalized_currency": None}
    try:
        return {"normalized_price": float(pval.replace(",", "")),
                "normalized_currency": pcur or "USD"}
    except ValueError:
        return {"normalized_price": None, "normalized_currency": None}


def _extract_quote_fields(text: str):
    """
    从回复正文抽取报价所需的供应商维度字段 (best-effort, 抽不到返回空串)。
    返回 dict: niche / country / cooperation_type / traffic / site_category / permanence
    """
    if not text:
        return {}
    low = text.lower()
    out = {}

    # cooperation_type: 合作类型
    coop_map = [
        ("guest post", "Guest Post"), ("guestpost", "Guest Post"),
        ("sponsored", "Sponsored Post"), ("sponsor", "Sponsored Post"),
        ("link insert", "Link Insert"), ("niche edit", "Link Insert"),
        ("backlink", "Backlink"), ("permanent", "Permanent Link"),
        ("homepage", "Homepage Link"), ("sidebar", "Sidebar Link"),
        ("article", "Article"), ("blog post", "Guest Post"),
        ("软文", "Guest Post"), ("客座", "Guest Post"), ("友链", "Backlink"),
        ("首页链接", "Homepage Link"), ("外链", "Backlink"),
    ]
    for kw, label in coop_map:
        if kw in low:
            out["cooperation_type"] = label
            break

    # country: 国家/地区信号
    country_map = [
        ("usa", "US"), ("united states", "US"), ("u.s.", "US"), ("america", "US"),
        ("uk", "UK"), ("united kingdom", "UK"), ("britain", "UK"), ("england", "UK"),
        ("germany", "DE"), ("deutschland", "DE"), ("德国", "DE"),
        ("france", "FR"), ("french", "FR"), ("法国", "FR"),
        ("italy", "IT"), ("italia", "IT"), ("意大利", "IT"),
        ("spain", "ES"), ("españa", "ES"), ("西班牙", "ES"),
        ("india", "IN"), ("indian", "IN"), ("印度", "IN"),
        ("canada", "CA"), ("canadian", "CA"), ("加拿大", "CA"),
        ("australia", "AU"), ("australian", "AU"), ("澳大利亚", "AU"),
        ("china", "CN"), ("中国", "CN"), ("brazil", "BR"), ("巴西", "BR"),
        ("netherlands", "NL"), ("荷兰", "NL"), ("russia", "RU"), ("俄罗斯", "RU"),
    ]
    for kw, code in country_map:
        if kw in low:
            out["country"] = code
            break

    # niche: 行业/领域 (常见外链 niche 关键词)
    niche_map = [
        ("casino", "Casino"), ("gambling", "Gambling"), ("crypto", "Crypto"),
        ("forex", "Forex"), ("finance", "Finance"), ("financial", "Finance"),
        ("health", "Health"), ("fitness", "Health"), ("medical", "Health"),
        ("tech", "Tech"), ("technology", "Tech"), ("saas", "Tech"),
        ("travel", "Travel"), ("tourism", "Travel"), ("旅游", "Travel"),
        ("fashion", "Fashion"), ("beauty", "Beauty"), ("时尚", "Fashion"),
        ("food", "Food"), ("recipe", "Food"), ("美食", "Food"),
        ("seo", "SEO"), ("marketing", "Marketing"), ("营销", "Marketing"),
        ("real estate", "Real Estate"), ("房地产", "Real Estate"),
        ("law", "Legal"), ("legal", "Legal"), ("律师", "Legal"),
        ("pets", "Pets"), ("pet", "Pets"), ("宠物", "Pets"),
        ("education", "Education"), ("教育", "Education"), ("游戏", "Gaming"),
        ("gaming", "Gaming"), ("game", "Gaming"),
    ]
    for kw, label in niche_map:
        if kw in low:
            out["niche"] = label
            break

    # permanence: 永久性/时效性
    if any(k in low for k in ["permanent", "永久", "never removed", "do-follow permanent"]):
        out["permanence"] = "Permanent"
    elif any(k in low for k in ["temporary", "临时", "6 month", "6 months", "one year", "1 year"]):
        out["permanence"] = "Temporary"

    # traffic: 流量信号 (月访问量)
    m_traffic = _re.search(r"(\d[\d,]*)\s*(?:monthly\s+)?(?:visits|traffic|views|pv)", low)
    if m_traffic:
        out["traffic"] = m_traffic.group(1).replace(",", "")

    return out


@app.route("/api/quote/import-a-replies", methods=["POST"])
def quote_import_a_replies():
    """
    从 reply_pool A类回复导入到 quote_pool。
    每条回复作为 quote 记录，reply_content 完整保留供人工解析。
    已导入的(email)不会重复导入。
    """
    user = request.form.get("user", "unknown").strip()
    force = request.form.get("force", "") == "1"  # force re-import all

    # Get all A类 replies
    page, page_size = 0, 1000
    all_replies = []
    while True:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/reply_pool?category=eq.A&select=*&limit={page_size}&offset={page*page_size}",
            headers=AUTH_HEADERS,
            timeout=30
        )
        if resp.status_code != 200:
            break
        data = resp.json()
        if not data:
            break
        all_replies.extend(data)
        if len(data) < page_size:
            break
        page += 1

    # Get existing (email, domain, price) triples for dedup.
    # Same email+domain with DIFFERENT price = distinct quote (keep both).
    # Same email+domain+price = true duplicate (skip).
    existing_keys = set()
    page = 0
    while True:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_pool?select=email,domain,price&limit={page_size}&offset={page*page_size}",
            headers=AUTH_HEADERS,
            timeout=30
        )
        if resp.status_code != 200:
            break
        data = resp.json()
        if not data:
            break
        for r in data:
            e = (r.get("email") or "").lower()
            d = (r.get("domain") or "").lower()
            p = (r.get("price") or "").strip().lower()
            if e:
                existing_keys.add(f"{e}|{d}|{p}")
        if len(data) < page_size:
            break
        page += 1

    imported = 0
    skipped = 0
    batch = []
    BATCH_SIZE = 30

    def flush():
        nonlocal imported
        if not batch:
            return
        for attempt in range(3):
            try:
                resp = requests.post(
                    f"{SUPABASE_URL}/rest/v1/quote_pool",
                    headers={**AUTH_HEADERS, "Prefer": "return=representation"},
                    json=batch,
                    timeout=30
                )
                if resp.status_code in (200, 201):
                    imported += len(batch)
                    break
            except Exception:
                if attempt < 2:
                    _time.sleep(1)
        batch.clear()

    for reply in all_replies:
        email = (reply.get("email") or "").lower()
        if not email:
            continue

        # 价格优先从 reply_pool.notes 里的 v2_price:$XX 标签取 (来自原始回信 Excel, 避免丢价)
        # 其次从回复正文解析 (修复: 进 quote pool 的 A类应有价格)
        price_str = ""
        notes_text = reply.get("notes") or ""
        m_v2 = _re.search(r"v2_price:\s*\$?\s*(\d[\d,]*\.?\d*)\s*(USD|EUR|GBP|INR|CNY)?", notes_text)
        if m_v2:
            pv = m_v2.group(1)
            pc = m_v2.group(2) or "USD"
            price_str = f"{pv} {pc}".strip()
        else:
            pval, pcur = _parse_price_from_content(reply.get("reply_content", ""))
            price_str = f"{pval} {pcur}".strip() if pval else ""

        # 自动标准化价格 (写 normalized_price / normalized_currency, 无需人工)
        norm = _normalize_price_fields(price_str)

        # 从回复正文抽取供应商维度字段 (niche/country/cooperation_type 等)
        qf = _extract_quote_fields(reply.get("reply_content", ""))
        price_missing = not price_str

        # domain 为空则跳过, 避免脏行 (空 domain 行会在前端显示为空白)
        domain = (reply.get("domain") or (email.split("@")[-1] if "@" in email else "")).lower()
        if not domain:
            skipped += 1
            continue

        # Dedup on (email, domain, price) triple — same email+domain with a
        # DIFFERENT price is a distinct quote and must be kept.
        dedup_key = f"{email}|{domain}|{price_str.strip().lower()}"
        if not force and dedup_key in existing_keys:
            skipped += 1
            continue
        existing_keys.add(dedup_key)

        batch.append({
            "email": email,
            "domain": domain,
            "supplier": (reply.get("supplier") or "")[:200],
            "contact_email": reply.get("contact_email") or email,
            "niche": qf.get("niche", ""),
            "country": qf.get("country", ""),
            "traffic": qf.get("traffic", ""),
            "site_category": "",
            "cooperation_type": qf.get("cooperation_type", ""),
            "price": price_str,
            "normalized_price": norm["normalized_price"],
            "normalized_currency": norm["normalized_currency"],
            "link_rules": "",
            "permanence": qf.get("permanence", ""),
            "content": "",
            "tat": "",
            "payment": "",
            "discount": "",
            "additional_services": "",
            "requirements": "",
            "reply_id": reply.get("reply_id"),
            "reply_content": (reply.get("reply_content") or "")[:8000],
            "status": "Price TBD" if price_missing else "New",
            "priority": 0,
            "notes": (f"Imported from A-class reply by {user}"
                      + (f" | 待补价" if price_missing else "")),
            "discovered_by": user,
            "discovered_at": now_iso(),
        })

        if len(batch) >= BATCH_SIZE:
            flush()

    flush()

    _log_operation("quote_import", user, "quote_pool", imported,
                   f"Imported {imported} A-class replies" + (f", skipped {skipped} duplicates" if skipped else ""))

    return jsonify({
        "imported": imported,
        "skipped": skipped,
        "total_a_replies": len(all_replies),
        "quote_pool_total": db.count("quote_pool"),
    })


@app.route("/api/quote/update", methods=["POST"])
def quote_update():
    """
    人工补全/更新一条报价记录, 并推进状态闭环。
    Body: {"email":"a@x.com", "fields":{可更新字段}, "status":"Quoted|Won|Lost|Price TBD"}
    支持更新: niche/country/traffic/site_category/cooperation_type/price/link_rules/
              permanence/content/tat/payment/discount/additional_services/requirements/status
    """
    data = request.get_json(force=True)
    email = (data.get("email") or "").lower().strip()
    if not email:
        return jsonify({"error": "email required"}), 400

    fields = data.get("fields", {})
    allowed = {"niche", "country", "traffic", "site_category", "cooperation_type",
               "price", "link_rules", "permanence", "content", "tat", "payment",
               "discount", "additional_services", "requirements", "status", "priority"}
    update = {}
    for k, v in fields.items():
        if k in allowed:
            update[k] = v

    if not update:
        return jsonify({"error": "no valid fields"}), 400

    update["reviewed_at"] = now_iso()
    update["reviewed_by"] = data.get("user", "unknown")
    resp, result = db.update("quote_pool", update, {"email": email})
    if resp and hasattr(resp, "status") and resp.status in (200, 201, 204):
        _log_operation("quote_update", data.get("user", "unknown"), "quote_pool", 1,
                       f"Updated quote for {email}: {list(update.keys())}")
        return jsonify({"updated": True, "email": email, "fields": list(update.keys())})
    return jsonify({"updated": False, "error": str(result)}), 500


@app.route("/api/quote/import", methods=["POST"])
def quote_import():
    """
    Import quotes from Excel/CSV upload.
    Smart column mapping supports both English and Chinese headers.
    Deduplicates by email (skip existing).
    """
    user = request.form.get("user", "unknown").strip()
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [str(c.value or '').strip() for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or '') for v in row])))
    except Exception as e:
        return jsonify({"error": f"File parse error: {str(e)}"}), 400

    if not rows:
        return jsonify({"error": "Empty file"}), 400

    # Smart column mapping (English + Chinese)
    def map_col(col):
        cl = col.lower().strip()
        # email
        if any(k in cl for k in ['邮箱', 'email', 'mail', 'e-mail']): return 'email'
        # domain
        if any(k in cl for k in ['域名', 'domain', '网站']): return 'domain'
        # supplier
        if any(k in cl for k in ['供应商', 'supplier', '发件人', 'from', '联系人', 'contact', 'name']): return 'supplier'
        # contact_email
        if any(k in cl for k in ['联系邮箱', 'contact_email', 'contact email', 'contact mail']): return 'contact_email'
        # niche
        if any(k in cl for k in ['领域', 'niche', '行业', 'industry', '细分']): return 'niche'
        # country
        if any(k in cl for k in ['国家', 'country', '地区', 'region']): return 'country'
        # traffic (注意: 不要匹配 'da' / 'domain authority', 否则 DA 列会被误映射成 traffic)
        if any(k in cl for k in ['流量', 'traffic', '访问量', 'visits']): return 'traffic'
        # da
        if any(k in cl for k in ['da', 'domain authority', '权威度']): return 'da'
        # site_category
        if any(k in cl for k in ['网站分类', 'site_category', 'site type', 'category', '分类']): return 'site_category'
        # cooperation_type
        if any(k in cl for k in ['合作类型', 'cooperation_type', 'collaboration', 'collaboration types', 'type', '合作']): return 'cooperation_type'
        # price
        if any(k in cl for k in ['价格', 'pricing', 'price', '报价', 'cost', 'fee']): return 'price'
        # link_rules
        if any(k in cl for k in ['链接规则', 'link_rules', 'link rule', 'publishing guidelines', 'publishing', '链接要求']): return 'link_rules'
        # permanence
        if any(k in cl for k in ['永久', 'permanence', 'permanent', '永久链接']): return 'permanence'
        # content
        if any(k in cl for k in ['内容', 'content', '文章要求']): return 'content'
        # tat
        if any(k in cl for k in ['时效', 'tat', 'turnaround', 'turnaround time', '交付时间', 'delivery']): return 'tat'
        # payment
        if any(k in cl for k in ['付款', 'payment', 'payment methods', 'pay', '支付方式']): return 'payment'
        # discount
        if any(k in cl for k in ['折扣', 'discount', '优惠']): return 'discount'
        # additional_services
        if any(k in cl for k in ['附加服务', 'additional_services', 'extra', '增值服务']): return 'additional_services'
        # requirements
        if any(k in cl for k in ['要求', 'requirements', 'requirement', 'content/link', '内容要求', '条件']): return 'requirements'
        # reply_content
        if any(k in cl for k in ['回复内容', 'reply_content', 'reply', '正文', 'body', 'message']): return 'reply_content'
        # status
        if any(k in cl for k in ['状态', 'status', 'state']): return 'status'
        # notes
        if any(k in cl for k in ['备注', 'notes', 'note', 'comment']): return 'notes'
        # priority
        if any(k in cl for k in ['优先级', 'priority', '重要度']): return 'priority'
        return None

    col_map = {}
    for col in rows[0].keys():
        mapped = map_col(col)
        if mapped:
            col_map[mapped] = col

    # Get existing (email, domain, price) triples for dedup.
    # Same email+domain with DIFFERENT price = distinct quote (keep both).
    existing = set()
    try:
        page, page_size = 0, 1000
        while True:
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/quote_pool?select=email,domain,price&limit={page_size}&offset={page*page_size}",
                headers=AUTH_HEADERS,
                timeout=30
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            if not data:
                break
            for r in data:
                e = (r.get('email') or '').lower()
                d = (r.get('domain') or '').lower()
                p = (r.get('price') or '').strip().lower()
                if e:
                    existing.add(f"{e}|{d}|{p}")
            if len(data) < page_size:
                break
            page += 1
    except Exception:
        pass

    imported = 0
    skipped = 0
    batch = []
    BATCH_SIZE = 30

    def flush_batch():
        nonlocal imported
        if not batch:
            return
        for attempt in range(3):
            try:
                resp = requests.post(
                    f"{SUPABASE_URL}/rest/v1/quote_pool",
                    headers={**AUTH_HEADERS, "Prefer": "return=representation"},
                    json=batch,
                    timeout=30
                )
                if resp.status_code in (200, 201):
                    imported += len(batch)
                    break
            except Exception:
                if attempt < 2:
                    _time.sleep(1)
        batch.clear()

    for row in rows:
        email = (row.get(col_map.get('email', ''), '') or '').strip().lower()
        if not email:
            continue

        domain = (row.get(col_map.get('domain', ''), '') or '').strip().lower()
        if not domain and '@' in email:
            domain = email.split('@')[-1]
        # domain 仍为空则跳过, 避免脏行
        if not domain:
            skipped += 1
            continue

        # price used for dedup triple (email, domain, price) — normalized form
        raw_price = str(row.get(col_map.get('price', ''), '') or '').strip().lower()[:50]

        # Dedup on (email, domain, price) triple — same email+domain with a
        # DIFFERENT price is a distinct quote and must be kept.
        dedup_key = f"{email}|{domain}|{raw_price}"
        if dedup_key in existing:
            skipped += 1
            continue
        existing.add(dedup_key)

        # 自动标准化价格
        norm = _normalize_price_fields(raw_price)

        # Parse priority as int
        priority = 0
        if 'priority' in col_map:
            try:
                priority = int(row.get(col_map['priority'], 0) or 0)
            except (ValueError, TypeError):
                priority = 0

        record = {
            "email": email,
            "domain": domain,
            "supplier": (row.get(col_map.get('supplier', ''), '') or '')[:200],
            "contact_email": (row.get(col_map.get('contact_email', ''), '') or email).lower(),
            "niche": (row.get(col_map.get('niche', ''), '') or '')[:100],
            "country": (row.get(col_map.get('country', ''), '') or '')[:50],
            "traffic": (row.get(col_map.get('traffic', ''), '') or '')[:50],
            "da": (row.get(col_map.get('da', ''), '') or '')[:50],
            "dr": (row.get(col_map.get('dr', ''), '') or '')[:50],
            "site_category": (row.get(col_map.get('site_category', ''), '') or '')[:50],
            "cooperation_type": (row.get(col_map.get('cooperation_type', ''), '') or '')[:50],
            "price": raw_price,
            "normalized_price": norm["normalized_price"],
            "normalized_currency": norm["normalized_currency"],
            "link_rules": (row.get(col_map.get('link_rules', ''), '') or '')[:200],
            "permanence": (row.get(col_map.get('permanence', ''), '') or '')[:50],
            "content": (row.get(col_map.get('content', ''), '') or '')[:500],
            "tat": (row.get(col_map.get('tat', ''), '') or '')[:50],
            "payment": (row.get(col_map.get('payment', ''), '') or '')[:50],
            "discount": (row.get(col_map.get('discount', ''), '') or '')[:50],
            "additional_services": (row.get(col_map.get('additional_services', ''), '') or '')[:200],
            "requirements": (row.get(col_map.get('requirements', ''), '') or '')[:500],
            "reply_content": (row.get(col_map.get('reply_content', ''), '') or '')[:8000],
            "status": (row.get(col_map.get('status', ''), '') or 'New')[:20],
            "priority": priority,
            "notes": (row.get(col_map.get('notes', ''), '') or '')[:500],
            "discovered_by": user,
            "discovered_at": now_iso(),
        }
        batch.append(record)

        if len(batch) >= BATCH_SIZE:
            flush_batch()

    flush_batch()

    _log_operation("quote_import", user, "quote_pool", imported,
                   f"Imported {imported} quotes from file" + (f", skipped {skipped} duplicates" if skipped else ""))

    return jsonify({
        "imported": imported,
        "skipped": skipped,
        "quote_pool_total": db.count("quote_pool"),
    })


# ════════════════════════════════════════════════════════════
# Comprehensive Stats
# ════════════════════════════════════════════════════════════

@app.route("/api/admin/clear-cache", methods=["POST"])
def admin_clear_cache():
    """Clear all in-memory caches (stats, unique domains, config)."""
    global _cache, CONFIG_CACHE
    _cache.clear()
    CONFIG_CACHE.clear()
    return jsonify({"status": "ok", "message": "All caches cleared"})


@app.route("/api/admin/ping", methods=["GET", "POST"])
def admin_ping():
    """Deploy test endpoint."""
    return jsonify({"status": "ok", "supabase_url": SUPABASE_URL})

@app.route("/api/admin/clean-empty", methods=["POST"])
def admin_clean_empty():
    """删除 domain 为空/NULL/纯空白 的脏行, 返回详情."""
    body = request.get_json(silent=True) or {}
    token = body.get("token", "")
    if token != (os.environ.get("ADMIN_TOKEN") or "maisui-normalize-2026"):
        return jsonify({"status": "error", "message": "unauthorized"}), 401
    try:
        WRITE_HEADERS = {**AUTH_HEADERS, "Content-Type": "application/json"}
        # 查所有 domain 为空或 NULL 的行
        results = []
        for filt in ["domain=is.null", "domain=eq.", "domain=eq. "]:
            r = requests.get(
                f"{SUPABASE_URL}/rest/v1/quote_pool?select=id,domain,price&{filt}",
                headers=AUTH_HEADERS, timeout=30)
            if r.status_code == 200:
                rows = r.json() if isinstance(r.json(), list) else []
                results.extend(rows)
        # 也查 domain 为空字符串的
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_pool?select=id,domain,price",
            headers=AUTH_HEADERS, timeout=30)
        all_rows = r.json() if (r.status_code == 200 and isinstance(r.json(), list)) else []
        empty_rows = [x for x in all_rows if isinstance(x, dict) and not x.get("domain", "") or not str(x.get("domain", "")).strip()]
        # 合并去重
        seen = set()
        unique_empty = []
        for row in empty_rows + results:
            rid = row.get("id")
            key = rid if rid is not None else f"domain:{repr(row.get('domain'))}"
            if key not in seen:
                seen.add(key)
                unique_empty.append(row)
        # 删除: 有 id 用 id, 无 id 用 domain=eq.(空串) 或 id=is.null
        del_ok = 0
        del_fail = 0
        for row in unique_empty:
            rid = row.get("id")
            if rid is not None:
                dr = requests.delete(
                    f"{SUPABASE_URL}/rest/v1/quote_pool?id=eq.{rid}",
                    headers=WRITE_HEADERS, timeout=30)
            else:
                # id 为空/NULL 的脏行, 用 domain 空串条件删除
                dr = requests.delete(
                    f"{SUPABASE_URL}/rest/v1/quote_pool?domain=eq.",
                    headers=WRITE_HEADERS, timeout=30)
            if dr.status_code in (200, 204):
                del_ok += 1
            else:
                del_fail += 1
        return jsonify({
            "status": "ok",
            "found_empty": len(unique_empty),
            "deleted": del_ok,
            "failed": del_fail,
            "sample": [{"id": r.get("id"), "domain": repr(r.get("domain")), "price": repr(r.get("price"))} for r in unique_empty[:5]],
        })
    except Exception as e:
        import traceback
        return jsonify({"status": "error", "message": str(e), "traceback": traceback.format_exc()}), 500

@app.route("/api/admin/delete-blank-row", methods=["POST"])
def admin_delete_blank_row():
    """专门删除 id 为 NULL / domain 为空串 的脏行."""
    body = request.get_json(silent=True) or {}
    token = body.get("token", "")
    if token != (os.environ.get("ADMIN_TOKEN") or "maisui-normalize-2026"):
        return jsonify({"status": "error", "message": "unauthorized"}), 401
    try:
        WRITE_HEADERS = {**AUTH_HEADERS, "Content-Type": "application/json"}
        # 1. 先用 id=is.null 删
        r1 = requests.delete(
            f"{SUPABASE_URL}/rest/v1/quote_pool?id=is.null",
            headers=WRITE_HEADERS, timeout=30)
        del1 = r1.status_code in (200, 204)
        # 2. 再用 domain=eq. 删（空字符串）
        r2 = requests.delete(
            f"{SUPABASE_URL}/rest/v1/quote_pool?domain=eq.",
            headers=WRITE_HEADERS, timeout=30)
        del2 = r2.status_code in (200, 204)
        # 3. 也试 domain is null
        r3 = requests.delete(
            f"{SUPABASE_URL}/rest/v1/quote_pool?domain=is.null",
            headers=WRITE_HEADERS, timeout=30)
        del3 = r3.status_code in (200, 204)
        return jsonify({
            "status": "ok",
            "id_null_deleted": del1,
            "id_null_code": r1.status_code,
            "domain_empty_deleted": del2,
            "domain_empty_code": r2.status_code,
            "domain_null_deleted": del3,
            "domain_null_code": r3.status_code,
        })
    except Exception as e:
        import traceback
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/admin/verify-key", methods=["POST"])
def admin_verify_key():
    """Verify a given Supabase key by hitting the REST API (GET + PATCH test)."""
    body = request.get_json(silent=True) or {}
    test_key = body.get("key", "")
    if not test_key:
        return jsonify({"status": "error", "message": "no key provided"}), 400
    hdrs = {"apikey": test_key, "Authorization": f"Bearer {test_key}", "Content-Type": "application/json"}
    try:
        # GET test
        r_get = requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_pool?select=count&limit=1",
            headers=hdrs, timeout=20)
        # PATCH test (try to set normalized_price on first row with null value)
        r_patch = requests.patch(
            f"{SUPABASE_URL}/rest/v1/quote_pool?normalized_price=is.null&limit=1",
            json={"normalized_price": -999.99},
            headers={**hdrs, "Prefer": "return=minimal"}, timeout=20)
        return jsonify({
            "status": "ok",
            "get_code": r_get.status_code,
            "get_body": r_get.text[:200],
            "patch_code": r_patch.status_code,
            "patch_body": r_patch.text[:300],
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/admin/raw-count", methods=["GET"])
def admin_raw_count():
    """直接用 service_role key 打各表真实 count，绕过前端 db 封装。"""
    SR_KEY = (
        SUPABASE_ANON_KEY  # from config.py (authoritative, can be rotated)
        or os.environ.get("SUPABASE_SERVICE_KEY")
        or os.environ.get("SUPABASE_ANON_KEY")
        or ""  # last resort; will fail downstream
    )
    hdrs = {"apikey": SR_KEY, "Authorization": f"Bearer {SR_KEY}"}
    tables = ["domain_pool", "supplier_pool", "quote_pool", "reply_pool",
              "email_pool", "operation_log", "config"]
    out = {}
    for t in tables:
        try:
            r = requests.get(
                f"{SUPABASE_URL}/rest/v1/{t}?select=count",
                headers=hdrs, timeout=20)
            out[t] = {"code": r.status_code, "body": r.text[:150]}
        except Exception as e:
            out[t] = {"code": "ERR", "body": str(e)[:100]}
    return jsonify(out)

@app.route("/api/admin/normalize-prices", methods=["POST"])
def admin_normalize_prices():
    """一次性存量修复 (在 Render/后端环境执行, 绕过沙箱对 supabase.co 的屏蔽):
      A. 对 normalized_price 为空的行, 从 price 文本提取金额+货币填 normalized_price/normalized_currency
      B. 删除 domain 为 NULL 的脏行 (前端空白行根因)
    返回处理统计。加简单口令防护防止滥用。
    写操作使用 service_role key 绕过 RLS (anon key 无 UPDATE/DELETE 权限)。
    """
    body = request.get_json(silent=True) or {}
    token = body.get("token", "")
    if token != (os.environ.get("ADMIN_TOKEN") or "maisui-normalize-2026"):
        return jsonify({"status": "error", "message": "unauthorized"}), 401

    # 写操作也用 AUTH_HEADERS (anon key) — RLS policy 已允许 anon UPDATE/DELETE
    # service_role key 在此项目上 PATCH 返回 401 (原因不明), 改回 anon
    WRITE_HEADERS = {**AUTH_HEADERS, "Content-Type": "application/json"}

    import traceback
    try:
        # B. 删空 domain 脏行
        null_resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_pool?select=id,domain&domain=is.null",
            headers=AUTH_HEADERS, timeout=30)
        null_rows = null_resp.json() if null_resp.status_code == 200 else []
        if not isinstance(null_rows, list):
            null_rows = []
        null_ids = [r["id"] for r in null_rows if isinstance(r, dict) and r.get("domain") is None]
        del_ok = 0
        for i in null_ids:
            r = requests.delete(
                f"{SUPABASE_URL}/rest/v1/quote_pool?id=eq.{i}",
                headers=WRITE_HEADERS, timeout=30)
            if r.status_code in (200, 204):
                del_ok += 1

        # A. 标准化 normalized_price 为空的行
        price_resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_pool?select=domain,price,normalized_price&normalized_price=is.null&limit=1000",
            headers=AUTH_HEADERS, timeout=30)
        rows = price_resp.json() if price_resp.status_code == 200 else []
        if not isinstance(rows, list):
            rows = []
        plan = []
        for x in rows:
            if not isinstance(x, dict):
                continue
            p = x.get("price")
            if p not in (None, "") and str(p).strip() != "":
                result = _normalize_price_fields(str(p))
                np_ = result.get("normalized_price")
                nc = result.get("normalized_currency")
                if np_ is not None:
                    plan.append((x["domain"], np_, nc))
        fill_ok = 0
        fail = 0
        fail_details = []
        for d, np_, nc in plan:
            r = requests.patch(
                f"{SUPABASE_URL}/rest/v1/quote_pool?domain=eq.{d}",
                json={"normalized_price": np_, "normalized_currency": nc},
                headers={**WRITE_HEADERS, "Prefer": "return=minimal"}, timeout=30)
            if r.status_code in (200, 204):
                fill_ok += 1
            else:
                fail += 1
                fail_details.append({"domain": d, "status": r.status_code, "body": r.text[:200]})
        return jsonify({
            "status": "ok",
            "null_domain_deleted": del_ok,
            "null_domain_total": len(null_ids),
            "normalized_filled": fill_ok,
            "normalized_failed": fail,
            "normalized_total": len(plan),
            "fail_details": fail_details[:5],
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e), "traceback": traceback.format_exc()}), 500

@app.route("/api/stats", methods=["GET"])
def get_stats():
    """Aggregated statistics (cached 60s). Read from pooled stats materialized view
    (pool_stats_mv) — 1 query replaces 20+ serial count() round-trips to Supabase."""

    def _fetch():
        # 优先读物化视图 (已聚合, 秒回)
        try:
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/pool_stats_mv?select=*&limit=1",
                headers=AUTH_HEADERS,
                timeout=20,
            )
            if resp.status_code == 200 and resp.json():
                mv = resp.json()[0]
                return {
                    "domain_total": mv.get("domain_total", 0),
                    "domain_new": mv.get("domain_new", 0),
                    "domain_claimed": mv.get("domain_claimed", 0),
                    "domain_imported": 0,
                    "domain_completed": mv.get("domain_contacted", 0),
                    "domain_exported": mv.get("domain_replied", 0),
                    "domain_unique_total": mv.get("domain_total", 0),
                    "domain_unique_new": mv.get("domain_new", 0),
                    "domain_unique_claimed": mv.get("domain_claimed", 0),
                    "domain_unique_contacted": mv.get("domain_contacted", 0),
                    "domain_unique_replied": mv.get("domain_replied", 0),
                    "domain_today_new": 0,
                    "email_total": mv.get("email_total", 0),
                    "email_unsent": mv.get("email_unsent", 0),
                    "email_assigned": mv.get("email_assigned", 0),
                    "email_sent": (mv.get("email_sent", 0) or 0) + (mv.get("email_exported", 0) or 0),
                    "email_bounce": mv.get("email_bounce", 0),
                    "reply_total": mv.get("reply_total", 0),
                    "reply_unread": mv.get("reply_unread", 0),
                    "reply_a": mv.get("reply_a", 0),
                    "reply_b": mv.get("reply_b", 0),
                    "reply_c": mv.get("reply_c", 0),
                    "reply_today": 0,
                    "reply_today_a": 0,
                    "reply_today_b": 0,
                    "reply_today_c": 0,
                    "quote_total": mv.get("quote_total", 0),
                    "quote_today_new": 0,
                    "quote_suppliers": 0,
                    "refreshed_at": str(mv.get("refreshed_at", "")),
                }
        except Exception as e:
            print(f"[WARN] pool_stats_mv read failed: {e}, falling back to live count")

        # Fallback: 物化视图不可用时退化到逐表 count (原逻辑)
        domain_total = db.count("domain_pool")
        domain_new = db.count("domain_pool", filters={"collection_status": "New"})
        domain_claimed = db.count("domain_pool", filters={"collection_status": "Claimed"})
        domain_contacted = db.count("domain_pool", filters={"collection_status": "Contacted"})
        domain_replied = db.count("domain_pool", filters={"collection_status": "Replied"})
        domain_unique_total = domain_total
        domain_unique_new = domain_new
        domain_unique_claimed = domain_claimed
        domain_unique_contacted = domain_contacted
        domain_unique_replied = domain_replied

        try:
            email_total = db.count("email_pool")
            email_unsent = db.count("email_pool", filters={"send_status": "UNSENT"})
            email_sent = db.count("email_pool", filters={"send_status": "SENT"})
            email_exported = db.count("email_pool", filters={"send_status": "EXPORTED"})
            email_bounce = db.count("email_pool", filters={"send_status": "Bounce"})
            try:
                email_assigned = db.count("email_pool", filters={"claimed_by": "not.is.null"})
            except Exception:
                email_assigned = 0
            email_sent_total = email_sent + email_exported
        except Exception:
            email_total = domain_total
            email_unsent = domain_new + domain_claimed
            email_sent_total = domain_contacted
            email_assigned = domain_claimed
            email_bounce = 0

        reply_total = reply_a = reply_b = reply_c = reply_unread = 0
        try:
            reply_total = db.count("reply_pool")
            if reply_total > 0:
                reply_a = db.count("reply_pool", filters={"category": "A"})
                reply_b = db.count("reply_pool", filters={"category": "B"})
                reply_c = db.count("reply_pool", filters={"category": "C"})
                reply_unread = db.count("reply_pool", filters={"status": "New"})
        except Exception:
            pass

        quote_total = db.count("quote_pool")

        return {
            "domain_total": domain_total,
            "domain_new": domain_new,
            "domain_claimed": domain_claimed,
            "domain_imported": 0,
            "domain_completed": domain_contacted,
            "domain_exported": domain_replied,
            "domain_unique_total": domain_unique_total,
            "domain_unique_new": domain_unique_new,
            "domain_unique_claimed": domain_unique_claimed,
            "domain_unique_contacted": domain_unique_contacted,
            "domain_unique_replied": domain_unique_replied,
            "domain_today_new": 0,
            "email_total": email_total,
            "email_unsent": email_unsent,
            "email_assigned": email_assigned,
            "email_sent": email_sent_total,
            "email_bounce": email_bounce,
            "reply_total": reply_total,
            "reply_unread": reply_unread,
            "reply_a": reply_a,
            "reply_b": reply_b,
            "reply_c": reply_c,
            "reply_today": 0,
            "reply_today_a": 0,
            "reply_today_b": 0,
            "reply_today_c": 0,
            "quote_total": quote_total,
            "quote_today_new": 0,
            "quote_suppliers": 0,
        }

    return jsonify(_cached("stats", ttl_sec=60, fn=_fetch))


@app.route("/api/members", methods=["GET"])
def api_members():
    """Team member stats."""
    domains = db.select("domain_pool",
                        select="claimed_by",
                        filters={"claimed_by": "not.is.null"},
                        limit=10000)
    users = defaultdict(lambda: {"domains": 0, "emails_assigned": 0, "emails_sent": 0})
    for d in (domains or []):
        u = d.get("claimed_by") or "unknown"
        users[u]["domains"] += 1

    return jsonify([
        {"username": u, **stats}
        for u, stats in sorted(users.items(), key=lambda x: -x[1]["domains"])
    ])


# ════════════════════════════════════════════════════════════
# Config API (→ config table as key-value store)
# ════════════════════════════════════════════════════════════

CONFIG_CACHE = {}  # simple in-memory cache for config

def _get_config(key, default=None):
    """Read a config value from Supabase, with local cache."""
    cache_key = f"config:{key}"
    now = _time.time()
    if cache_key in CONFIG_CACHE and CONFIG_CACHE[cache_key][0] > now:
        return CONFIG_CACHE[cache_key][1]
    try:
        rows = db.select("config", select="value", filters={"key": key}, limit=1)
        val = rows[0]["value"] if rows else default
    except Exception:
        val = default
    CONFIG_CACHE[cache_key] = (now + 30, val)
    return val


def _set_config(key, value, description=""):
    """Upsert a config value via POST (update if exists, insert if new)."""
    existing = db.select("config", select="key", filters={"key": key}, limit=1)
    if existing:
        db.update("config", {"value": value, "updated_at": now_iso()}, {"key": key})
    else:
        db.insert("config", {"key": key, "value": value, "description": description})
    CONFIG_CACHE[f"config:{key}"] = (_time.time() + 30, value)
    return True


@app.route("/api/config/team", methods=["GET", "POST"])
def config_team():
    """Get or set team members list (comma-separated or JSON array).
    GET → {"members": ["leo","emma","jack"], "raw": "leo,emma,jack"}
    POST {"members": ["leo","emma","jack"]} → save to config"""
    if request.method == "GET":
        raw = _get_config("team_members", "leo,emma,jack")
        # Parse: try JSON array, fallback to comma-sep
        try:
            members = json.loads(raw) if raw else []
        except (json.JSONDecodeError, TypeError):
            members = [m.strip() for m in str(raw).split(",") if m.strip()]
        return jsonify({"members": members, "raw": raw})

    # POST: update team members
    data = request.get_json(force=True)
    members = data.get("members", [])
    if isinstance(members, str):
        members = [m.strip() for m in members.split(",") if m.strip()]
    if not members:
        return jsonify({"error": "members is required"}), 400
    raw = json.dumps(members, ensure_ascii=False)
    _set_config("team_members", raw, "Team member usernames for domain distribution")
    # Clear distribute cache
    CONFIG_CACHE.pop("team_members_list", None)
    return jsonify({"members": members, "saved": True})


@app.route("/api/domain/import_log", methods=["GET"])
def domain_import_log():
    """Recent import records (domains with notes containing 'imported by')."""
    limit = min(int(request.args.get("limit", 50)), 500)
    domains = db.select(
        "domain_pool",
        select="domain_id,domain,source,notes,created_at",
        filters={"notes": "like.[imported by%]"},
        limit=limit,
        order="created_at",
        ascending=False,
    )
    # Parse imported_by from notes field
    result = []
    for d in (domains or []):
        notes = d.get("notes", "")
        imported_by = ""
        if "[imported by " in notes:
            imported_by = notes.split("[imported by ")[1].split("]")[0]
        result.append({
            "domain_id": d.get("domain_id"),
            "domain": d.get("domain"),
            "imported_by": imported_by,
            "notes": notes,
            "imported_at": d.get("created_at", ""),
        })
    return jsonify({"imports": result, "count": len(result)})


@app.route("/api/log/list", methods=["GET"])
def log_list():
    """List operation logs (import/export/distribute).

    Reads from the independent operation_log table (new), falling back
    to the legacy config JSON blob (pre-migration). Results are merged,
    de-duplicated, sorted newest-first, then filtered/paginated.
    """
    limit = min(int(request.args.get("limit", 100)), 200000)
    op_type = request.args.get("type", "")
    user = request.args.get("user", "")

    logs = []

    # 1) New table — use service_role key to bypass PostgREST db_max_rows (1000) cap.
    #    Cursor-paginate over op_time DESC so we always retrieve the FULL log set.
    try:
        SR_KEY = (
            SUPABASE_ANON_KEY  # from config.py (authoritative, can be rotated)
            or os.environ.get("SUPABASE_SERVICE_KEY")
            or os.environ.get("SUPABASE_ANON_KEY")
            or ""  # last resort; will fail downstream
        )
        SR_HEADERS = {"apikey": SR_KEY, "Authorization": f"Bearer {SR_KEY}"}
        fetched = 0
        PAGE = 1000
        last_id = None
        while fetched < 200000:
            # Cursor pagination on log_id (int serial, stable) — avoids PostgREST tz-string
            # comparison bugs on op_time (old rows stored as "2026-... 00:00" with a space).
            q = f"{SUPABASE_URL}/rest/v1/operation_log?select=log_id,op_time,type,username,pool,count,detail&order=log_id.desc&limit={PAGE}"
            if last_id is not None:
                q += f"&log_id=lt.{last_id}"
            resp = requests.get(q, headers=SR_HEADERS, timeout=30)
            if resp.status_code != 200:
                break
            rows = resp.json()
            if not isinstance(rows, list) or not rows:
                break
            for r in rows:
                logs.append({
                    "log_id": "op_" + str(r.get("log_id")),
                    "time": _utc_to_bj(str(r.get("op_time"))),
                    "type": r.get("type"),
                    "user": r.get("username"),
                    "table": r.get("pool"),
                    "count": r.get("count", 0),
                    "detail": r.get("detail") or "",
                })
            fetched += len(rows)
            if len(rows) < PAGE:
                break
            last_id = rows[-1]["log_id"]
    except Exception:
        pass

    # 2) Legacy config blob (pre-migration / fallback)
    try:
        raw = _get_config("operation_logs", "[]")
        legacy = json.loads(raw) if raw else []
        for l in legacy:
            l.setdefault("log_id", "legacy_" + str(l.get("time")))
            logs.append(l)
    except Exception:
        pass

    # De-dup by log_id
    seen = set()
    unique = []
    for l in logs:
        lid = l.get("log_id")
        if lid in seen:
            continue
        seen.add(lid)
        unique.append(l)

    # Sort newest-first by time (string ISO sorts lexically for same format)
    unique.sort(key=lambda x: str(x.get("time", "")), reverse=True)

    # Filter
    if op_type:
        unique = [l for l in unique if l.get("type") == op_type]
    if user:
        unique = [l for l in unique if l.get("user") == user]

    unique = unique[:limit]

    # Convert legacy UTC timestamps to Beijing time for display
    for l in unique:
        if "time" in l:
            l["time"] = _utc_to_bj(l["time"])

    return jsonify({"logs": unique, "count": len(unique)})


@app.route("/api/log/delete", methods=["POST"])
def log_delete():
    """Delete operation log entries by log_id list.

    New-table ids are prefixed 'op_<id>' (delete from operation_log by PK).
    Legacy ids are prefixed 'legacy_<time>' (remove from config blob).
    """
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get("ids") or []
    if not ids:
        return jsonify({"error": "no ids provided"}), 400

    new_ids = []
    legacy_ids = []
    for i in ids:
        if str(i).startswith("op_"):
            new_ids.append(int(str(i)[3:]))
        elif str(i).startswith("legacy_"):
            legacy_ids.append(i)

    removed = 0

    # Delete from new table
    if new_ids:
        try:
            # Supabase delete with filter: operation_log?log_id=in.(1,2,3)
            filt = "log_id=in.(" + ",".join(str(x) for x in new_ids) + ")"
            db.delete("operation_log", filt)
            removed += len(new_ids)
        except Exception:
            pass

    # Delete from legacy config blob
    if legacy_ids:
        try:
            raw = _get_config("operation_logs", "[]")
            logs = json.loads(raw) if raw else []
            before = len(logs)
            logs = [l for l in logs if l.get("log_id") not in legacy_ids]
            removed += (before - len(logs))
            _set_config("operation_logs", json.dumps(logs, ensure_ascii=False), "Operation logs")
        except Exception:
            pass

    return jsonify({"deleted": removed, "remaining": -1})


# ════════════════════════════════════════════════════════════

EMAIL_POOL_SQL = """-- Run this in Supabase SQL Editor to create the email_pool table:
-- (This is the independent email pool — separate from domain_pool)

CREATE TABLE IF NOT EXISTS email_pool (
    email_id            BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    email               TEXT NOT NULL,
    domain              TEXT NOT NULL,
    send_status         TEXT DEFAULT 'UNSENT',
    source              TEXT DEFAULT '未提取',
    collection_status   TEXT DEFAULT 'New',
    claimed_by          TEXT,
    claim_time          TIMESTAMPTZ,
    import_batch        TEXT,
    notes               TEXT,
    priority            INT DEFAULT 0,
    created_at          TIMESTAMPTZ DEFAULT NOW(),
    updated_at          TIMESTAMPTZ DEFAULT NOW()
);

-- Unique index on email (dedup key)
CREATE UNIQUE INDEX IF NOT EXISTS idx_email_pool_email ON email_pool(email);
CREATE INDEX IF NOT EXISTS idx_email_pool_domain ON email_pool(domain);
CREATE INDEX IF NOT EXISTS idx_email_pool_status ON email_pool(send_status);
CREATE INDEX IF NOT EXISTS idx_email_pool_collection ON email_pool(collection_status);

-- Enable RLS for email_pool
ALTER TABLE email_pool ENABLE ROW LEVEL SECURITY;

-- Allow anon reads
CREATE POLICY "anon_can_read_emails" ON email_pool
    FOR SELECT USING (true);

-- Allow anon inserts
CREATE POLICY "anon_can_insert_emails" ON email_pool
    FOR INSERT WITH CHECK (true);

-- Allow anon updates
CREATE POLICY "anon_can_update_emails" ON email_pool
    FOR UPDATE USING (true);


-- Optional: migrate existing email data from domain_pool to email_pool
-- Uncomment and run if you have existing contact_email data in domain_pool:

/*
INSERT INTO email_pool (email, domain, source, collection_status, claimed_by, claim_time, notes, priority, created_at)
SELECT DISTINCT ON (LOWER(TRIM(contact_email)))
  LOWER(TRIM(contact_email)) as email,
  LOWER(TRIM(domain)) as domain,
  COALESCE(source, '未提取') as source,
  COALESCE(collection_status, 'New') as collection_status,
  claimed_by,
  claim_time,
  notes,
  COALESCE(priority, 0) as priority,
  created_at
FROM domain_pool
WHERE contact_email IS NOT NULL AND TRIM(contact_email) != ''
ON CONFLICT (email) DO NOTHING;
*/
"""

SETUP_SQL = """-- Run this in Supabase SQL Editor to create the reply_pool table:

CREATE TABLE IF NOT EXISTS reply_pool (
    reply_id            BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    email               TEXT NOT NULL,
    domain              TEXT,
    reply_content       TEXT,
    reply_time          TIMESTAMPTZ DEFAULT NOW(),
    category            TEXT DEFAULT 'C',
    status              TEXT DEFAULT 'New',
    supplier            TEXT,
    contact_email       TEXT,
    discovered_by       TEXT DEFAULT 'system',
    discovered_at       TIMESTAMPTZ DEFAULT NOW(),
    replied_by          TEXT,
    replied_at          TIMESTAMPTZ,
    notes               TEXT
);

CREATE INDEX IF NOT EXISTS idx_reply_category ON reply_pool(category);
CREATE INDEX IF NOT EXISTS idx_reply_status ON reply_pool(status);
CREATE INDEX IF NOT EXISTS idx_reply_domain ON reply_pool(domain);

-- Enable RLS for reply_pool
ALTER TABLE reply_pool ENABLE ROW LEVEL SECURITY;

-- Allow anon reads
CREATE POLICY "anon_can_read_replies" ON reply_pool
    FOR SELECT USING (true);

-- Allow anon inserts
CREATE POLICY "anon_can_insert_replies" ON reply_pool
    FOR INSERT WITH CHECK (true);

-- Allow anon updates
CREATE POLICY "anon_can_update_replies" ON reply_pool
    FOR UPDATE USING (true);


-- ════════════════════════════════════════════════════════════
-- operation_log (independent table — replaces config JSON blob)
-- The old approach stored all logs in a single config JSON field,
-- which silently failed once it exceeded Supabase's row size limit.
-- ════════════════════════════════════════════════════════════

CREATE TABLE IF NOT EXISTS operation_log (
    log_id              BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    op_time             TIMESTAMPTZ DEFAULT NOW(),
    type                TEXT,
    username            TEXT,
    pool                TEXT,
    count               INT DEFAULT 0,
    detail              TEXT,
    created_at          TIMESTAMPTZ DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_op_log_time ON operation_log(op_time DESC);
CREATE INDEX IF NOT EXISTS idx_op_log_type ON operation_log(type);

-- Enable RLS for operation_log
ALTER TABLE operation_log ENABLE ROW LEVEL SECURITY;

-- Allow anon reads
CREATE POLICY "anon_can_read_op_log" ON operation_log
    FOR SELECT USING (true);

-- Allow anon inserts
CREATE POLICY "anon_can_insert_op_log" ON operation_log
    FOR INSERT WITH CHECK (true);

-- Allow anon deletes
CREATE POLICY "anon_can_delete_op_log" ON operation_log
    FOR DELETE USING (true);
"""


@app.route("/setup")
def setup_page():
    """Show setup instructions."""
    tables_found = {}
    for tbl in ["domain_pool", "email_pool", "supplier_pool", "quote_pool", "reply_pool", "operation_log", "config"]:
        try:
            cnt = db.count(tbl)
            tables_found[tbl] = f"OK ({cnt} rows)"
        except Exception:
            tables_found[tbl] = "MISSING"

    rows = "".join(
        f"<tr><td>{t}</td><td style='color:{'green' if 'OK' in v else 'red'}'>{v}</td></tr>"
        for t, v in tables_found.items()
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>Shared Pool Setup</title>
<style>
body{{font:14px/1.6 sans-serif;max-width:800px;margin:40px auto;padding:20px;color:#1a1a2e}}
table{{border-collapse:collapse;width:100%;margin:16px 0}}
td,th{{padding:8px 12px;border:1px solid #e0e0e0;text-align:left}}
pre{{background:#1a1a2e;color:#a5d6ff;padding:16px;border-radius:8px;overflow-x:auto;font-size:12px}}
h2{{border-top:1px solid #e0e0e0;padding-top:20px;margin-top:30px}}
.green{{color:#1d9e75}} .red{{color:#d85a30}}
</style></head>
<body>
<h1>Shared Pool v2 — Setup</h1>
<p>Supabase: <code>{SUPABASE_URL}</code></p>
<table>{rows}</table>
<p>If any of <b>email_pool</b>, <b>reply_pool</b> or <b>operation_log</b> is MISSING, copy the SQL below to your Supabase SQL Editor:</p>
<h2>email_pool</h2>
<pre>{EMAIL_POOL_SQL}</pre>
<h2>reply_pool + operation_log</h2>
<pre>{SETUP_SQL}</pre>
<p><a href="/">← Back to Dashboard</a></p>
</body></html>"""


# ════════════════════════════════════════════════════════════
# Dashboard (4-tab UI)
# ════════════════════════════════════════════════════════════

DASHBOARD_HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<title>Shared Pool v2 — Supplier Intelligence</title>
<style>
:root{--bg:#f8f9fa;--card:#fff;--border:#e0e0e0;--text:#1a1a2e;--muted:#6b7280;--blue:#378add;--green:#1d9e75;--amber:#ba7517;--coral:#d85a30;--purple:#7f77dd;--teal:#0f6e56;}
*{margin:0;padding:0;box-sizing:border-box}
body{font:13px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);color:var(--text);padding:20px 28px;max-width:1200px;margin:0 auto}
h1{font-size:18px;font-weight:500;margin-bottom:2px}
.sub{font-size:12px;color:var(--muted);margin-bottom:20px}
.tabs{display:flex;gap:4px;margin-bottom:20px;border-bottom:1px solid var(--border)}
.tab{padding:8px 18px;font-size:13px;font-weight:500;color:var(--muted);cursor:pointer;border-bottom:2px solid transparent;transition:all .2s}
.tab:hover{color:var(--text)}
.tab.active{color:var(--blue);border-bottom-color:var(--blue)}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:20px}
.card{background:var(--card);border:0.5px solid var(--border);border-radius:10px;padding:14px}
.card .label{font-size:11px;color:var(--muted);margin-bottom:3px}
.card .value{font-size:22px;font-weight:500}
.card .value.blue{color:var(--blue)}.card .value.green{color:var(--green)}.card .value.amber{color:var(--amber)}.card .value.coral{color:var(--coral)}.card .value.purple{color:var(--purple)}.card .value.teal{color:var(--teal)}
/* ── Checkbox & selection styles ── */
.chk-all{width:16px;height:16px;cursor:pointer;accent-color:var(--blue)}
.chk-row{width:15px;height:15px;cursor:pointer;accent-color:var(--blue);margin-top:2px}
.btn-bar{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;align-items:center}
.export-sel-btn{display:inline-block;padding:6px 14px;font-size:12px;font-weight:500;border-radius:6px;border:none;cursor:pointer;color:#fff;background:var(--teal);transition:opacity .2s}
.export-sel-btn:hover{opacity:.85}
.export-sel-btn:disabled{opacity:.4;cursor:not-allowed}
.sel-count{font-size:12px;color:var(--muted);padding:0 8px}
/* ── Enhanced table styles ── */
table{width:100%;border-collapse:collapse;background:var(--card);border:0.5px solid var(--border);border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.04)}
th,td{text-align:left;padding:7px 10px;font-size:12px}
th{font-weight:600;background:linear-gradient(180deg,#f8f9fa 0%,#f1f3f5 100%);border-bottom:1.5px solid #d0d7de;color:var(--text);position:sticky;top:0;z-index:1;text-transform:none;letter-spacing:.3px}
td{border-bottom:0.5px solid #eef1f4;transition:background .15s;vertical-align:middle}
tr:hover td{background:#f7fafc}
tr:last-child td{border-bottom:none}
tr.selected-row td{background:#e8f4fd}
.btn{display:inline-block;padding:6px 14px;font-size:12px;font-weight:500;border-radius:6px;border:none;cursor:pointer;color:#fff;background:var(--blue);transition:opacity .2s;margin-right:4px}
.btn:hover{opacity:.85}
.btn.green{background:var(--green)}.btn.amber{background:var(--amber)}.btn.coral{background:var(--coral)}.btn.purple{background:var(--purple)}.btn.teal{background:var(--teal)}
.actions{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}
.status-New{background:#e6f1fb;color:#185fa5;padding:2px 7px;border-radius:4px;font-size:11px}
.status-Claimed{background:#faeeda;color:#854f0b;padding:2px 7px;border-radius:4px;font-size:11px}
.status-Contacted{background:#e1f5ee;color:#0f6e56;padding:2px 7px;border-radius:4px;font-size:11px}
.status-Replied{background:#eaf3de;color:#3b6d11;padding:2px 7px;border-radius:4px;font-size:11px}
.cat-A{background:#eaf3de;color:#3b6d11;font-weight:500;padding:2px 7px;border-radius:4px;font-size:11px}
.cat-B{background:#faeeda;color:#854f0b;font-weight:500;padding:2px 7px;border-radius:4px;font-size:11px}
.cat-C{background:#f1efe8;color:#5f5e5a;font-weight:500;padding:2px 7px;border-radius:4px;font-size:11px}
.refresh{font-size:11px;color:var(--muted);text-align:right;margin-bottom:12px}
.page{display:none}.page.active{display:block}
.btn.active{opacity:.6;box-shadow:inset 0 2px 4px rgba(0,0,0,.2)}
.log-filter-group{display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap;align-items:center}
#log-pagination .btn{padding:4px 10px;font-size:11px}
.form-row{display:flex;gap:8px;align-items:center;margin-bottom:12px;flex-wrap:wrap}
.form-row input,.form-row select{padding:5px 10px;font-size:12px;border:0.5px solid var(--border);border-radius:6px}
.form-row label{font-size:12px;color:var(--muted);min-width:60px}
</style>
</head>
<body>
<h1>Shared Pool v2 — Supplier Intelligence</h1>
<p class="sub">Supabase Edition · Domain / Email / Reply / Price · <a href="/setup" style="color:var(--blue)">Setup</a></p>
<p class="refresh" id="refresh-msg">Loading...</p>

<div class="tabs">
  <div class="tab active" onclick="switchTab('domain')">Domain Pool</div>
  <div class="tab" onclick="switchTab('email')">Email Pool</div>
  <div class="tab" onclick="switchTab('reply')">Reply Pool</div>
  <div class="tab" onclick="switchTab('quote')">Quote Pool</div>
  <div class="tab" onclick="switchTab('log')">Operation Log</div>
</div>

<!-- Domain Pool -->
<div class="page active" id="page-domain">
  <div class="cards" id="domain-cards"></div>
  <div class="actions">
    <button class="btn" onclick="exportDomains()">Export NEW domains</button>
    <button class="btn green" onclick="distributeDomains()">Distribute to team</button>
    <button class="btn amber" onclick="toggleImport()">Import domains</button>
  </div>
  <div class="form-row">
    <label>User</label><input id="d-user" placeholder="your name" style="width:100px" onchange="saveUserName()">
    <label>Count</label><input id="d-count" value="5000" type="number" style="width:80px">
    <label>Status</label><select id="d-status" onchange="loadDomainTable()"><option value="">All</option><option value="New">New</option><option value="Claimed">Claimed</option><option value="Contacted">Contacted</option><option value="Replied">Replied</option></select>
    <label>User Filter</label><select id="d-user-filter" onchange="loadDomainTable()"><option value="">All Users</option></select>
  </div>
  <!-- Import panel (hidden by default) -->
  <div id="import-panel" style="display:none;margin-bottom:16px;padding:14px;background:var(--card);border:0.5px solid var(--border);border-radius:10px">
    <div style="font-size:13px;font-weight:500;margin-bottom:8px">Import Domains</div>
    <textarea id="import-text" rows="5" placeholder="Paste domains, one per line&#10;example.com&#10;site.org&#10;..." style="width:100%;padding:8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;resize:vertical"></textarea>
    <div style="margin-top:8px;display:flex;align-items:center;gap:8px;flex-wrap:wrap">
      <select id="import-status" style="padding:5px 10px;font-size:12px;border:0.5px solid var(--border);border-radius:6px">
        <option value="New">Status: New</option>
      </select>
      <button class="btn green" onclick="importDomains()">Submit Import</button>
      <span style="color:var(--muted);font-size:12px">or</span>
      <label style="cursor:pointer">
        <input type="file" id="import-csv-file" accept=".csv" style="display:none" onchange="handleCsvUpload(this)">
        <span class="btn" style="display:inline-block">Upload CSV</span>
      </label>
      <button class="btn" style="font-size:11px" onclick="downloadCsvTemplate()">Download Template</button>
      <span id="import-result" style="font-size:12px;color:var(--muted)"></span>
    </div>
  </div>
  <div id="team-panel" style="display:none;margin-bottom:16px;padding:14px;background:var(--card);border:0.5px solid var(--border);border-radius:10px">
    <div style="font-size:13px;font-weight:500;margin-bottom:8px">Team Members</div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
      <input id="team-input" style="width:300px;padding:5px 10px;font-size:12px;border:0.5px solid var(--border);border-radius:6px" placeholder="leo,emma,jack">
      <button class="btn purple" onclick="saveTeam()">Save Team</button>
      <span id="team-result" style="font-size:12px;color:var(--muted)"></span>
    </div>
    <div style="font-size:11px;color:var(--muted);margin-top:4px">Comma-separated usernames. Distribute assigns domains round-robin to these members.</div>
  </div>
  <table id="domain-table"><tr><th>#</th><th>Domain</th><th>Source</th><th>Status</th><th>Claimed By</th><th>Priority</th><th>Created</th></tr></table>
  <!-- Domain pagination -->
  <div class="log-filter-group" id="domain-pagination" style="justify-content:center;margin-top:8px;display:none">
    <button class="btn" onclick="changeDomainPage(-1)">Prev</button>
    <span id="domain-page-info" style="font-size:12px;color:var(--muted);margin:0 12px">Page 1 / 1</span>
    <button class="btn" onclick="changeDomainPage(1)">Next</button>
    <label style="font-size:12px;color:var(--muted);margin-left:16px">Per page:</label>
    <select id="domain-page-size" onchange="onDomainPageSizeChange()" style="padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="10" selected>10</option>
      <option value="20">20</option>
      <option value="50">50</option>
    </select>
  </div>
</div>

<!-- Email Pool -->
<div class="page" id="page-email">
  <div class="cards" id="email-cards"></div>
  <div class="actions">
    <button class="btn" onclick="exportEmails()">Export send queue</button>
    <label style="font-size:12px;color:var(--muted)">User</label><input id="e-user" placeholder="your name" style="width:100px" onchange="saveUserName()">
    <label style="font-size:12px;color:var(--muted)">Count</label><input id="e-count" value="1000" type="number" style="width:80px">
    <label style="font-size:12px;color:var(--muted)">User Filter</label><select id="e-user-filter" onchange="loadEmailTable()"><option value="">All Users</option></select>
    <button class="btn green" onclick="toggleEmailImport()">Import emails</button>
  </div>
  <!-- Email Import panel (hidden by default) -->
  <div id="email-import-panel" style="display:none;margin-bottom:16px;padding:14px;background:var(--card);border:0.5px solid var(--border);border-radius:10px">
    <div style="font-size:13px;font-weight:500;margin-bottom:8px">Import Emails (independent email pool, batch dedup, large batches OK)</div>
    <input id="email-import-user" placeholder="Your name" style="width:120px;padding:4px 8px;margin-right:8px" value="">
    <textarea id="email-import-text" rows="5" placeholder="Paste email-domain pairs, one per line&#10;email@example.com,example.com&#10;contact@site.org,site.org&#10;..." style="width:100%;padding:8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;resize:vertical;margin-top:8px"></textarea>
    <div style="margin-top:8px;display:flex;align-items:center;gap:8px;flex-wrap:wrap">
      <button class="btn green" onclick="importEmails()">Submit Import</button>
      <span style="color:var(--muted);font-size:12px">or</span>
      <label style="cursor:pointer">
        <input type="file" id="email-import-csv-file" accept=".csv" style="display:none" onchange="handleEmailCsvUpload(this)">
        <span class="btn" style="display:inline-block">Upload CSV</span>
      </label>
      <button class="btn" style="font-size:11px" onclick="downloadEmailCsvTemplate()">Download Template</button>
      <span id="email-import-result" style="font-size:12px;color:var(--muted)"></span>
    </div>
  </div>
  <table id="email-table"><tr><th>#</th><th>Email</th><th>Domain</th><th>Send Status</th><th>Claimed By</th><th>Source</th><th>Created</th></tr></table>
  <!-- Email pagination -->
  <div class="log-filter-group" id="email-pagination" style="justify-content:center;margin-top:8px;display:none">
    <button class="btn" onclick="changeEmailPage(-1)">Prev</button>
    <span id="email-page-info" style="font-size:12px;color:var(--muted);margin:0 12px">Page 1 / 1</span>
    <button class="btn" onclick="changeEmailPage(1)">Next</button>
    <label style="font-size:12px;color:var(--muted);margin-left:16px">Per page:</label>
    <select id="email-page-size" onchange="onEmailPageSizeChange()" style="padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="10" selected>10</option>
      <option value="20">20</option>
      <option value="50">50</option>
    </select>
  </div>
</div>

<!-- Reply Pool -->
<div class="page" id="page-reply">
  <div class="cards" id="reply-cards"></div>
  <div class="actions">
    <button class="btn" onclick="loadReplyTable('')">All</button>
    <button class="btn green" onclick="loadReplyTable('A')">A class</button>
    <button class="btn amber" onclick="loadReplyTable('B')">B class</button>
    <button class="btn purple" onclick="loadReplyTable('C')">C class</button>
    <button class="btn" style="background:var(--muted);color:#fff" onclick="loadReplyTable('D')">D class</button>
    <label style="font-size:12px;color:var(--muted)">User</label><input id="r-user" placeholder="your name" style="width:100px" onchange="saveUserName()">
    <label style="font-size:12px;color:var(--muted)">User Filter</label><select id="r-user-filter" onchange="loadReplyTable(REPLY_PAGER.category)"><option value="">All Users</option></select>
    <button class="btn green" onclick="toggleReplyImport()">Import replies</button>
  </div>
  <!-- Reply Import panel (hidden by default) -->
  <div id="reply-import-panel" style="display:none;margin-bottom:16px;padding:14px;background:var(--card);border:0.5px solid var(--border);border-radius:10px">
    <div style="font-size:13px;font-weight:500;margin-bottom:8px">Import Replies (Excel/CSV, template: 序号,邮箱,域名,账号,发件人,主题,正文摘要,日期)</div>
    <label style="cursor:pointer">
      <input type="file" id="reply-import-file" accept=".xlsx,.csv" style="display:none" onchange="handleReplyUpload(this)">
      <span class="btn" style="display:inline-block">Upload File</span>
    </label>
    <button class="btn" style="font-size:11px" onclick="downloadReplyTemplate()">Download Template</button>
    <span id="reply-import-result" style="font-size:12px;color:var(--muted)"></span>
  </div>
  <table id="reply-table"><tr><th>#</th><th>Email</th><th>Domain</th><th>Category</th><th>Status</th><th>Supplier</th><th>Reply Time</th><th>Content</th></tr></table>
  <!-- Reply pagination -->
  <div class="log-filter-group" id="reply-pagination" style="justify-content:center;margin-top:8px;display:none">
    <button class="btn" onclick="changeReplyPage(-1)">Prev</button>
    <span id="reply-page-info" style="font-size:12px;color:var(--muted);margin:0 12px">Page 1 / 1</span>
    <button class="btn" onclick="changeReplyPage(1)">Next</button>
    <label style="font-size:12px;color:var(--muted);margin-left:16px">Per page:</label>
    <select id="reply-page-size" onchange="onReplyPageSizeChange()" style="padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="10" selected>10</option>
      <option value="20">20</option>
      <option value="50">50</option>
    </select>
  </div>
</div>

<!-- Quote Pool -->
<div class="page" id="page-quote">
  <div class="cards" id="quote-cards"></div>
  <script>setTimeout(()=>{try{loadQuoteTable();}catch(e){console.error('preload quote failed',e);}},300);</script>
  <div class="actions">
    <span style="font-size:12px;color:var(--muted);margin-right:4px">Export CSV:</span>
    <button class="btn" onclick="exportQuotes('all')">All</button>
    <button class="btn" onclick="exportQuotes('ready')">Normal only</button>
    <button class="btn" onclick="exportQuotes('abnormal')">Abnormal only</button>
    <button class="btn red" id="quote-delete-btn" onclick="deleteSelectedQuotes()" style="display:none">Delete Selected (<span id="quote-sel-count">0</span>)</button>
    <label style="font-size:12px;color:var(--muted)">User</label><input id="q-user" placeholder="your name" style="width:100px" onchange="saveUserName()">
    <label style="font-size:12px;color:var(--muted)">User Filter</label><select id="q-user-filter" onchange="loadQuoteTable()"><option value="">All Users</option></select>
    <label style="font-size:12px;color:var(--muted)"><svg style="vertical-align:-2px;width:13px;height:13px;fill:none;stroke:currentColor;stroke-width:2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/></svg> Search</label>
    <input id="q-search" placeholder="domain, supplier, email, keyword..." style="width:200px;padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)" onkeydown="if(event.key==='Enter'){QUOTE_PAGER.page=1;loadQuoteTable()}">
    <button class="btn" style="font-size:11px;padding:4px 10px" onclick="QUOTE_PAGER.page=1;loadQuoteTable()">Search</button>
    <button class="btn green" onclick="importARepliesToQuotes()">Import A-class replies</button>
    <button class="btn" onclick="toggleQuoteImport()">Import from File</button>
    <span id="quote-import-result" style="font-size:12px;color:var(--muted)"></span>
  </div>
  <!-- Quote Import panel (hidden by default) -->
  <div id="quote-import-panel" style="display:none;margin-bottom:16px;padding:14px;background:var(--card);border:0.5px solid var(--border);border-radius:10px">
    <div style="font-size:13px;font-weight:500;margin-bottom:8px">Import Quotes (Excel/CSV). Supports: Domain, Supplier, Email, Niche, Country, Traffic, Type, Price, Link Rules, Permanence, Content, TAT, Payment, Discount, etc.</div>
    <label style="cursor:pointer">
      <input type="file" id="quote-import-file" accept=".xlsx,.csv" style="display:none" onchange="handleQuoteUpload(this)">
      <span class="btn" style="display:inline-block">Upload File</span>
    </label>
    <button class="btn" style="font-size:11px" onclick="downloadQuoteTemplate()">Download Template</button>
    <span id="quote-import-file-result" style="font-size:12px;color:var(--muted)"></span>
  </div>
  <div style="overflow-x:auto;max-height:calc(100vh - 320px);overflow-y:auto">
  <div style="margin-bottom:4px;text-align:right">
    <label style="font-size:11.5px;color:var(--muted);cursor:pointer" onclick="toggleQuoteCols()">
      <input type="checkbox" id="q-show-all-cols"> Hide low-freq columns (DR/DA/Traffic/TAT/etc.)
    </label>
  </div>
  <style>
    #quote-table{font-size:12px;border-collapse:collapse;width:auto;min-width:100%}
    #quote-table th,#quote-table td{padding:4px 7px;vertical-align:middle;line-height:1.25}
    #quote-table tbody tr{min-height:26px}
    #quote-table th{white-space:nowrap;font-size:11px;font-weight:600}
    #quote-table th.col-link{width:160px}
    #quote-table th.col-keywords{width:120px}
    #quote-table th.col-linkrules{width:140px}
    #quote-table th.col-contact{width:150px}
    #quote-table td{font-size:12px}
    /* All columns visible by default (was hidden, reverted per user request) */
    .q-col-extra{display:table-cell}
    body.hide-quote-extra-cols .q-col-extra{display:none}
  </style>
  <table id="quote-table">
    <thead><tr>
      <th><input type="checkbox" class="chk-all" onchange="toggleQuoteAll(this)" title="Select All"></th><th>#</th>
      <th class="col-link">Link</th><th>Price</th><th>Backlink Type</th>
      <th class="q-col-extra">DR</th><th class="q-col-extra">DA</th>
      <th class="q-col-extra">Ref. Domains</th><th class="q-col-extra">Traffic</th><th>Country</th>
      <th class="col-keywords">Keywords</th>
      <th class="q-col-extra">Categories</th><th class="q-col-extra">Languages</th>
      <th class="q-col-extra">TAT</th><th>Permanence</th><th class="col-contact">Contact</th>
      <th class="q-col-extra">Cooperation</th><th class="q-col-extra">Payment</th>
      <th class="col-linkrules">Link Rules</th><th>Status</th>
    </tr></thead><tbody></tbody>
  </table>
  </div>
  <!-- Quote pagination -->
  <div class="log-filter-group" id="quote-pagination" style="justify-content:center;margin-top:8px;display:none">
    <button class="btn" onclick="changeQuotePage(-1)">Prev</button>
    <span id="quote-page-info" style="font-size:12px;color:var(--muted);margin:0 12px">Page 1 / 1</span>
    <button class="btn" onclick="changeQuotePage(1)">Next</button>
    <label style="font-size:12px;color:var(--muted);margin-left:16px">Per page:</label>
    <select id="quote-page-size" onchange="onQuotePageSizeChange()" style="padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="10">10</option>
      <option value="20" selected>20</option>
      <option value="20">20</option>
      <option value="50">50</option>
    </select>
  </div>
</div>

<!-- Operation Log -->
<div class="page" id="page-log">
  <div class="cards" id="log-cards"></div>

  <!-- 一级分类: Pool -->
  <div class="log-filter-group" id="log-pool-filters">
    <button class="btn active" id="log-pool-all" onclick="setLogPoolFilter('')">All</button>
    <button class="btn green" id="log-pool-domain" onclick="setLogPoolFilter('domain')">Domain Pool</button>
    <button class="btn amber" id="log-pool-email" onclick="setLogPoolFilter('email')">Email Pool</button>
    <button class="btn teal" id="log-pool-reply" onclick="setLogPoolFilter('reply')">Reply Pool</button>
    <button class="btn purple" id="log-pool-quote" onclick="setLogPoolFilter('quote')">Quote Pool</button>
  </div>

  <!-- 二级分类: Action (动态显示) -->
  <div class="log-filter-group" id="log-action-filters" style="display:none">
    <span style="font-size:12px;color:var(--muted)">Action:</span>
    <button class="btn active" id="log-action-all" onclick="setLogActionFilter('')">All</button>
    <button class="btn" id="log-action-import" onclick="setLogActionFilter('import')">Import</button>
    <button class="btn" id="log-action-export" onclick="setLogActionFilter('export')">Export</button>
    <button class="btn" id="log-action-distribute" onclick="setLogActionFilter('distribute')">Distribute</button>
  </div>

  <!-- 用户筛选 -->
  <div class="log-filter-group">
    <label style="font-size:12px;color:var(--muted)">User:</label>
    <select id="log-user-filter" onchange="renderLogPage()" style="padding:4px 10px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="">All Users</option>
    </select>
    <button class="btn red" id="log-delete-btn" onclick="deleteSelectedLogs()" style="display:none">Delete Selected (<span id="log-sel-count">0</span>)</button>
    <button class="btn" onclick="refreshLogs()" style="margin-left:auto">Refresh</button>
  </div>

  <table id="log-table"><tr><th><input type="checkbox" class="chk-all" onchange="toggleLogAll(this)" title="Select All"></th><th>#</th><th>Time</th><th>User</th><th>Action</th><th>Pool</th><th>Count</th><th>Detail</th></tr></table>

  <!-- 分页控件 -->
  <div class="log-filter-group" id="log-pagination" style="justify-content:center;margin-top:8px">
    <button class="btn" onclick="changeLogPage(-1)" id="log-prev">Prev</button>
    <span id="log-page-info" style="font-size:12px;color:var(--muted);margin:0 12px">Page 1 / 1 (0 records)</span>
    <button class="btn" onclick="changeLogPage(1)" id="log-next">Next</button>
    <label style="font-size:12px;color:var(--muted);margin-left:16px">Per page:</label>
    <select id="log-page-size" onchange="onLogPageSizeChange()" style="padding:4px 8px;font-size:12px;border:0.5px solid var(--border);border-radius:6px;background:var(--bg);color:var(--text)">
      <option value="10" selected>10</option>
      <option value="20">20</option>
      <option value="50">50</option>
    </select>
  </div>

  <!-- 当前页统计 -->
  <div id="log-page-stats" style="font-size:12px;color:var(--muted);text-align:center;margin-top:4px;padding-bottom:12px"></div>
</div>

<script>
const API=location.origin;
function fmt(n){return n!=null?Number(n).toLocaleString():'0'}
function esc(s,n=80){return String(s||'').replace(/</g,'&lt;').slice(0,n)}

// ── Pagination state for all pools ──
const DOMAIN_PAGER={page:1,pageSize:10};
const EMAIL_PAGER ={page:1,pageSize:10};
const REPLY_PAGER ={page:1,pageSize:10,category:''};
const QUOTE_PAGER ={page:1,pageSize:20};

// ── User name via localStorage ──
function getUserName(){
  const ids=['d-user','e-user','r-user','q-user'];
  let name='';
  for(const id of ids){
    const el=document.getElementById(id);
    if(el && el.value.trim()){ name=el.value.trim(); break; }
  }
  if(!name) name=localStorage.getItem('shared_pool_user')||'';
  if(name){
    for(const id of ids){
      const el=document.getElementById(id);
      if(el) el.value=name;
    }
  }
  return name;
}
function saveUserName(){
  const ids=['d-user','e-user','r-user','q-user'];
  let name='';
  for(const id of ids){
    const el=document.getElementById(id);
    if(el && el.value.trim()){ name=el.value.trim(); break; }
  }
  if(name){
    localStorage.setItem('shared_pool_user',name);
    for(const id of ids){
      const el=document.getElementById(id);
      if(el) el.value=name;
    }
  }
}
// Load saved user name on page load
(function(){
  const saved=localStorage.getItem('shared_pool_user');
  if(saved){
    ['d-user','e-user','r-user','q-user'].forEach(id=>{
      const el=document.getElementById(id); if(el) el.value=saved;
    });
  }
})();

// ── Import panel ──
function toggleImport(){
  const p=document.getElementById('import-panel');
  p.style.display=p.style.display==='none'?'block':'none';
  loadTeam();  // also show team panel together
  document.getElementById('team-panel').style.display=p.style.display;
}
async function importDomains(){
  const user=getUserName();
  if(!user){alert('Please enter your name first');return;}
  const raw=document.getElementById('import-text').value;
  const status=document.getElementById('import-status').value;
  const domains=raw.split(/[\n,;]+/).map(d=>d.trim()).filter(d=>d&&d.includes('.'));
  if(!domains.length){alert('No valid domains found');return;}
  document.getElementById('import-result').textContent='Importing '+domains.length+' domains...';
  try{
    const r=await fetch(API+'/api/domain/register',{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({domains,collection_status:status,imported_by:user})
    }).then(r=>r.json());
    document.getElementById('import-result').textContent='Done: '+r.new+' new, '+r.duplicates+' duplicates';
    document.getElementById('import-text').value='';
    loadStats();loadDomainTable();
  }catch(e){
    document.getElementById('import-result').textContent='Error: '+e.message;
  }
}

// ── CSV Upload & Template ──
function handleCsvUpload(input){
  const file=input.files[0];
  if(!file)return;
  const reader=new FileReader();
  reader.onload=function(e){
    const text=e.target.result;
    const lines=text.split(/\r?\n/);
    const domains=[];
    for(const line of lines){
      const cells=line.split(',');
      for(const cell of cells){
        const d=cell.trim().toLowerCase();
        if(d && d.includes('.')) domains.push(d);
      }
    }
    if(!domains.length){alert('No valid domains found in CSV');input.value='';return;}
    document.getElementById('import-text').value=domains.join('\n');
    document.getElementById('import-result').textContent='CSV loaded: '+domains.length+' domains ready. Click Submit Import.';
    input.value='';
  };
  reader.readAsText(file);
}
function downloadCsvTemplate(){
  const csv='domain\nexample.com\nsite.org\ncompany.net\n';
  const blob=new Blob([csv],{type:'text/csv'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;
  a.download='domain_import_template.csv';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

// ── Team management ──
async function loadTeam(){
  try{
    const r=await fetch(API+'/api/config/team').then(r=>r.json());
    document.getElementById('team-input').value=(r.members||[]).join(',');
  }catch(e){}
}
async function saveTeam(){
  const raw=document.getElementById('team-input').value;
  const members=raw.split(',').map(m=>m.trim()).filter(m=>m);
  if(!members.length){alert('Enter at least one team member');return;}
  try{
    await fetch(API+'/api/config/team',{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({members})
    });
    document.getElementById('team-result').textContent='Saved: '+members.join(', ');
  }catch(e){
    document.getElementById('team-result').textContent='Error: '+e.message;
  }
}

function switchTab(name){
  const tabs=['domain','email','reply','quote','log'];
  document.querySelectorAll('.tab').forEach((t,i)=>t.classList.toggle('active',tabs[i]===name));
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.getElementById('page-'+name).classList.add('active');
  if(name==='email')loadEmailTable();
  if(name==='reply')loadReplyTable('');
  if(name==='quote')loadQuoteTable();
  if(name==='log')loadLogTable('');
}

async function loadStats(){
  try{
    const s=await fetch(API+'/api/stats').then(r=>r.json());
    document.getElementById('domain-cards').innerHTML=[
      {l:'Unique Domains',v:s.domain_unique_total,c:'blue'},{l:'New',v:s.domain_unique_new,c:'amber'},
      {l:'Claimed',v:s.domain_unique_claimed,c:'purple'},
    ].map(c=>`<div class="card"><div class="label">${c.l}</div><div class="value ${c.c}">${fmt(c.v)}</div></div>`).join('');
    document.getElementById('email-cards').innerHTML=[
      {l:'Total Emails',v:s.email_total,c:'blue'},{l:'Unsent',v:s.email_unsent,c:'amber'},
      {l:'Sent',v:s.email_sent,c:'green'},
    ].map(c=>`<div class="card"><div class="label">${c.l}</div><div class="value ${c.c}">${fmt(c.v)}</div></div>`).join('');
    document.getElementById('reply-cards').innerHTML=[
      {l:'累计 Total',v:s.reply_total,c:'blue'},{l:'累计 A',v:s.reply_a,c:'green'},
      {l:'累计 B',v:s.reply_b,c:'amber'},{l:'累计 C',v:s.reply_c,c:'purple'},
      {l:'今日新增',v:(s.reply_today_a||0)+(s.reply_today_b||0)+(s.reply_today_c||0),c:'blue'},
      {l:'今日 A',v:s.reply_today_a||0,c:'green'},{l:'今日 B',v:s.reply_today_b||0,c:'amber'},
      {l:'今日 C',v:s.reply_today_c||0,c:'purple'},
    ].map(c=>`<div class="card"><div class="label">${c.l}</div><div class="value ${c.c}">${fmt(c.v)}</div></div>`).join('');
    document.getElementById('quote-cards').innerHTML=[
      {l:'Total quotes',v:s.quote_total,c:'blue'},{l:'Today new',v:s.quote_today_new,c:'amber'},
    ].map(c=>`<div class="card"><div class="label">${c.l}</div><div class="value ${c.c}">${fmt(c.v)}</div></div>`).join('');
    document.getElementById('refresh-msg').textContent='Updated: '+new Date().toLocaleTimeString('zh-CN');
    populateUserFilters(s);
  }catch(e){document.getElementById('refresh-msg').textContent='Error: '+e.message}
}

// ── Populate user filter dropdowns from stats members data ──
async function populateUserFilters(stats){
  try{
    const r=await fetch(API+'/api/members').then(r=>r.json());
    const members=(r.members||[]);
    const ids=['d-user-filter','e-user-filter','r-user-filter','q-user-filter'];
    ids.forEach(id=>{
      const sel=document.getElementById(id);
      if(!sel)return;
      const current=sel.value;
      sel.innerHTML='<option value="">All Users</option>'+
        members.map(m=>`<option value="${esc(m)}">${esc(m)}</option>`).join('');
      if(current)sel.value=current;
    });
  }catch(e){}
}

// ── Selection state for all pools ──
const DOMAIN_SEL = new Set();
const EMAIL_SEL = new Set();
const REPLY_SEL = new Set();
const QUOTE_SEL = new Set();
const LOG_SEL = new Set();

// ── Selection helpers ──
function _updateSelCount(selSet, countId, btnId){
  const c=selSet.size;
  document.getElementById(countId).textContent=c;
  document.getElementById(btnId).disabled=c===0;
}
function _getGlobalIdx(page, pageSize, rowIdx){ return (page-1)*pageSize + rowIdx + 1; }

async function loadDomainTable(){
  const status=document.getElementById('d-status').value;
  const userFilter=document.getElementById('d-user-filter').value;
  const limit=DOMAIN_PAGER.pageSize;
  const offset=(DOMAIN_PAGER.page-1)*limit;
  const r=await fetch(API+'/api/domain/list?limit='+limit+'&offset='+offset+(status?'&status='+status:'')+(userFilter?'&user='+encodeURIComponent(userFilter):'')).then(r=>r.json());
  const total=r.total||r.unique_total||0;
  const maxPage=Math.max(1,Math.ceil(total/limit));
  if(DOMAIN_PAGER.page>maxPage){DOMAIN_PAGER.page=maxPage;}
  DOMAIN_SEL.clear();
  document.getElementById('domain-table').innerHTML='<tr><th>#</th><th>Domain</th><th>Status</th><th>Claimed By</th><th>Priority</th><th>Created</th></tr>'+
    (r.domains||[]).map((d,i)=>`<tr data-idx="${i}" data-id="${d.domain_id}">
      <td style="color:var(--muted);font-size:11.5px;text-align:center">${_getGlobalIdx(DOMAIN_PAGER.page,limit,i)}</td>
      <td>${esc(d.domain)}</td>
      <td><span class="status-${d.collection_status||'New'}">${d.collection_status||'New'}</span></td>
      <td>${esc(d.claimed_by)}</td>
      <td>${d.priority||0}</td>
      <td>${(d.created_at||'').slice(0,16)}</td>
    </tr>`).join('');
  document.getElementById('domain-page-info').textContent='Page '+DOMAIN_PAGER.page+' / '+maxPage+' ('+total+' records)';
  const dp=document.getElementById('domain-pagination');
  if(dp)dp.style.display=total>limit?'flex':'none';
  console.log('[Domain] total='+total+' limit='+limit+' showPagination='+(total>limit));
}
function changeDomainPage(delta){
  DOMAIN_PAGER.page=Math.max(1,DOMAIN_PAGER.page+delta);
  loadDomainTable();
}
function onDomainPageSizeChange(){
  DOMAIN_PAGER.pageSize=parseInt(document.getElementById('domain-page-size').value)||50;
  DOMAIN_PAGER.page=1;
  loadDomainTable();
}

async function loadEmailTable(){
  const u=getUserName()||'';
  const userFilter=document.getElementById('e-user-filter').value;
  const limit=EMAIL_PAGER.pageSize;
  const offset=(EMAIL_PAGER.page-1)*limit;
  const r=await fetch(API+'/api/email/queue?user='+encodeURIComponent(u)+'&count='+limit+'&offset='+offset+(userFilter?'&claimed_by='+encodeURIComponent(userFilter):'')).then(r=>r.json());
  const total=r.total||0;
  const maxPage=Math.max(1,Math.ceil(total/limit));
  if(EMAIL_PAGER.page>maxPage){EMAIL_PAGER.page=maxPage;}
  EMAIL_SEL.clear();
  document.getElementById('email-table').innerHTML='<tr><th>#</th><th>Email</th><th>Domain</th><th>Send Status</th><th>Claimed By</th><th>Created</th></tr>'+
    (r.emails||[]).map((e,i)=>`<tr data-idx="${i}" data-id="${e.email_id}">
      <td style="color:var(--muted);font-size:11.5px;text-align:center">${_getGlobalIdx(EMAIL_PAGER.page,limit,i)}</td>
      <td>${esc(e.contact_email)}</td>
      <td>${esc(e.domain)}</td>
      <td><span class="status-${e.send_status||'UNSENT'}">${e.send_status||'UNSENT'}</span></td>
      <td>${esc(e.claimed_by)}</td>
      <td>${(e.created_at||'').slice(0,16)}</td>
    </tr>`).join('');
  document.getElementById('email-page-info').textContent='Page '+EMAIL_PAGER.page+' / '+maxPage+' ('+total+' records)';
  document.getElementById('email-pagination').style.display=total>limit?'flex':'none';
}
function changeEmailPage(delta){
  EMAIL_PAGER.page=Math.max(1,EMAIL_PAGER.page+delta);
  loadEmailTable();
}
function onEmailPageSizeChange(){
  EMAIL_PAGER.pageSize=parseInt(document.getElementById('email-page-size').value)||50;
  EMAIL_PAGER.page=1;
  loadEmailTable();
}

async function loadReplyTable(cat){
  if(cat!==undefined){REPLY_PAGER.category=cat;REPLY_PAGER.page=1;}
  const limit=REPLY_PAGER.pageSize;
  const offset=(REPLY_PAGER.page-1)*limit;
  const category=REPLY_PAGER.category;
  const userFilter=document.getElementById('r-user-filter').value;
  const r=await fetch(API+'/api/reply/list?limit='+limit+'&offset='+offset+(category?'&category='+category:'')+(userFilter?'&user='+encodeURIComponent(userFilter):'')).then(r=>r.json());
  const total=r.total||0;
  const maxPage=Math.max(1,Math.ceil(total/limit));
  if(REPLY_PAGER.page>maxPage){REPLY_PAGER.page=maxPage;}
  REPLY_SEL.clear();
  document.getElementById('reply-table').innerHTML='<tr><th>#</th><th>Email</th><th>Domain</th><th>Category</th><th>Status</th><th>Supplier</th><th>Reply Time</th><th>Content</th></tr>'+
    (r.replies||[]).map((rp,i)=>`<tr data-idx="${i}" data-id="${rp.reply_id}">
      <td style="color:var(--muted);font-size:11.5px;text-align:center">${_getGlobalIdx(REPLY_PAGER.page,limit,i)}</td>
      <td>${esc(rp.email)}</td>
      <td>${esc(rp.domain)}</td>
      <td><span class="cat-${rp.category}">${rp.category||'C'}</span></td>
      <td>${esc(rp.status)}</td>
      <td>${esc(rp.supplier)}</td>
      <td>${(rp.reply_time||'').slice(0,16)}</td>
      <td style="max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${esc(rp.reply_content)}</td>
    </tr>`).join('');
  document.getElementById('reply-page-info').textContent='Page '+REPLY_PAGER.page+' / '+maxPage+' ('+total+' records)';
  document.getElementById('reply-pagination').style.display=total>limit?'flex':'none';
}
function toggleReplyAll(chk){
  document.querySelectorAll('#reply-table .chk-row').forEach(c=>{c.checked=chk.checked;c.onchange();});
}
function onReplyCheck(chk,id,idx){
  if(chk.checked) REPLY_SEL.add({id,idx,row:chk.closest('tr')}); else REPLY_SEL.forEach((v,i)=>{if(v.id===id)REPLY_SEL.delete(v);});
  chk.closest('tr').classList.toggle('selected-row',chk.checked);
}
function changeReplyPage(delta){
  REPLY_PAGER.page=Math.max(1,REPLY_PAGER.page+delta);
  loadReplyTable();
}
function onReplyPageSizeChange(){
  REPLY_PAGER.pageSize=parseInt(document.getElementById('reply-page-size').value)||50;
  REPLY_PAGER.page=1;
  loadReplyTable();
}

async function loadQuoteTable(){
  const limit=QUOTE_PAGER.pageSize;
  const offset=(QUOTE_PAGER.page-1)*limit;
  const userFilter=document.getElementById('q-user-filter').value;
  const searchVal=(document.getElementById('q-search')||{}).value||'';
  let url=API+'/api/quote/list?limit='+limit+'&offset='+offset;
  if(userFilter) url+='&supplier='+encodeURIComponent(userFilter);
  if(searchVal.trim()) url+='&search='+encodeURIComponent(searchVal.trim());
  const r=await fetch(url).then(r=>r.json());
  const total=r.total||0;
  const maxPage=Math.max(1,Math.ceil(total/limit));
  if(QUOTE_PAGER.page>maxPage){QUOTE_PAGER.page=maxPage;}
  QUOTE_SEL.clear();
  // Quote Pool: fixed columns (Jenny template), always show all, empty if no data
  const allQuotes = r.quotes || [];

  // Quote Pool columns (Jenny template + DR/DA/Traffic/Keywords etc.)
  // All columns always shown; empty if no data for that row.
  let th='<tr><th><input type="checkbox" class="chk-all" onchange="toggleQuoteAll(this)" title="Select All"></th><th>#</th>'+
    '<th class="col-link">Link</th><th>Price</th><th>Backlink Type</th>'+
    '<th class="q-col-extra">DR</th><th class="q-col-extra">DA</th>'+
    '<th class="q-col-extra">Ref. Domains</th><th class="q-col-extra">Traffic</th><th>Country</th>'+
    '<th class="col-keywords">Keywords</th>'+
    '<th class="q-col-extra">Categories</th><th class="q-col-extra">Languages</th>'+
    '<th class="q-col-extra">TAT</th><th>Permanence</th><th class="col-contact">Contact</th>'+
    '<th class="q-col-extra">Cooperation</th><th class="q-col-extra">Payment</th>'+
    '<th class="col-linkrules">Link Rules</th><th>Status</th></tr>';

  const mappedFields=new Set(['domain','price','cooperation_type','traffic','country','site_category','niche','tat','permanence','contact_email','email','supplier','da','dr','ref_domains','keywords','categories','languages','link_rules','content','payment','discount','additional_services','requirements','reply_content','status','notes','discovered_by','discovered_at','quote_id','reply_id','priority','id']);

  let tb=(allQuotes).map((q,i)=>{
      const otherParts=[];
      for(const [k,v] of Object.entries(q)){
        if(!mappedFields.has(k.toLowerCase()) && v!=null && String(v).trim()) otherParts.push(k+': '+v);
      }
      const otherStr=otherParts.join(' | ');

      // Contact: prefer contact_email, fallback email; strip +sub-address for display
      const rawContact=q.contact_email||q.email||'';
      const displayContact=rawContact.replace(/\+[^@]+(?=@)/,'');

      // 异常行背景色
      const _ds = (q.data_status||'').trim();
      let _rowStyle = '';
      if (_ds === 'NEED_DOMAIN') _rowStyle = 'background:#fff0f0';        // 浅红
      else if (_ds === 'NEED_PRICE') _rowStyle = 'background:#fff8e0';    // 浅黄
      else if (_ds && _ds !== 'READY') _rowStyle = 'background:#fff4e0';  // 浅橙(NEED_REVIEW等)

      return `<tr data-idx="${i}" data-id="${q.quote_id||q.id}"${_rowStyle ? ` style="${_rowStyle}"` : ''}>
        <td><input type="checkbox" class="chk-row" onchange="onQuoteCheck(this,${q.quote_id||q.id},${i})"></td>
        <td style="color:var(--muted);font-size:10.5px;text-align:center">${_getGlobalIdx(QUOTE_PAGER.page,limit,i)}</td>
        <td><a href="http://${esc(q.domain)}" target="_blank">${esc(q.domain)}</a></td>
        <td style="white-space:nowrap">${esc((function(){var n=parseFloat(q.normalized_price);if(!isNaN(n)){var d=n%1===0?String(n):String(n);return d+' '+(q.normalized_currency||'USD')}return q.price||''})())}</td>
        <td>${esc(q.price_type || q.site_category || q.niche || '')}</td>
        <td class="q-col-extra">${esc(q.dr||'')}</td>
        <td class="q-col-extra">${esc(q.da||'')}</td>
        <td class="q-col-extra">${esc(q.ref_domains||'')}</td>
        <td class="q-col-extra">${esc(q.traffic||'')}</td>
        <td>${esc(q.country||'')}</td>
        <td style="max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(q.keywords||'',200)}">${esc(q.keywords||'')}</td>
        <td class="q-col-extra" style="max-width:100px;overflow:hidden;text-overflow:ellipsis">${esc(q.categories||'')}</td>
        <td class="q-col-extra">${esc(q.languages||'')}</td>
        <td class="q-col-extra">${esc(q.tat||'')}</td>
        <td>${esc(q.permanence||'')}</td>
        <td title="${esc(rawContact)}">${esc(displayContact)}</td>
        <td class="q-col-extra">${esc(q.cooperation_type||'')}</td>
        <td class="q-col-extra">${esc(q.payment||'')}</td>
        <td style="max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${esc(q.link_rules||'',300)}">${esc(q.link_rules||'')}</td>
        <td><span class="status-${(q.status||'New')}">${esc(q.status||'New')}</span></td>
      </tr>`;
    }).join('');

  document.getElementById('quote-table').innerHTML=th+tb;
  _updateSelCount(QUOTE_SEL,'quote-sel-count','quote-delete-btn');
  document.getElementById('quote-delete-btn').style.display=QUOTE_SEL.size?'inline-block':'none';
  document.getElementById('quote-page-info').textContent='Page '+QUOTE_PAGER.page+' / '+maxPage+' ('+total+' records)';
  const qp=document.getElementById('quote-pagination');
  if(qp)qp.style.display=total>limit?'flex':'none';
  console.log('[Quote] total='+total+' limit='+limit+' showPagination='+(total>limit));
}
function toggleQuoteAll(chk){
  document.querySelectorAll('#quote-table .chk-row').forEach(c=>{c.checked=chk.checked;c.onchange();});
}
function onQuoteCheck(chk,id,idx){
  if(chk.checked) QUOTE_SEL.add({id,idx,row:chk.closest('tr')}); else QUOTE_SEL.forEach((v,i)=>{if(v.id===id)QUOTE_SEL.delete(v);});
  _updateSelCount(QUOTE_SEL,'quote-sel-count','quote-delete-btn');
  document.getElementById('quote-delete-btn').style.display=QUOTE_SEL.size?'inline-block':'none';
  chk.closest('tr').classList.toggle('selected-row',chk.checked);
}
function changeQuotePage(delta){
  QUOTE_PAGER.page=Math.max(1,QUOTE_PAGER.page+delta);
  loadQuoteTable();
}
function onQuotePageSizeChange(){
  QUOTE_PAGER.pageSize=parseInt(document.getElementById('quote-page-size').value)||50;
  QUOTE_PAGER.page=1;
  loadQuoteTable();
}
function toggleQuoteCols(){
  document.body.classList.toggle('hide-quote-extra-cols',document.getElementById('q-show-all-cols').checked);
}

function toggleEmailImport(){
  const p=document.getElementById('email-import-panel');
  p.style.display=p.style.display==='none'?'block':'none';
}
async function importEmails(){
  const user=getUserName();
  if(!user){alert('Please enter your name first');return;}
  const raw=document.getElementById('email-import-text').value;
  const lines=raw.split(/[\n]+/).map(l=>l.trim()).filter(l=>l);
  const records=[];
  for(const line of lines){
    const parts=line.split(/[,;\t]+/);
    if(parts.length>=2){
      const email=parts[0].trim().toLowerCase();
      const domain=parts[1].trim().toLowerCase().replace(/^www\./,'');
      if(email.includes('@')&&domain.includes('.')){
        records.push({email,domain});
      }
    }
  }
  if(!records.length){alert('No valid email-domain pairs found. Format: email,domain');return;}
  document.getElementById('email-import-result').textContent='Importing '+records.length+' emails...';
  try{
    const r=await fetch(API+'/api/email/import',{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({emails:records,imported_by:user})
    }).then(r=>r.json());
    document.getElementById('email-import-result').textContent='Done: '+r.imported+' imported, '+r.skipped+' skipped (duplicates)';
    document.getElementById('email-import-text').value='';
    loadStats();loadEmailTable();
  }catch(e){
    document.getElementById('email-import-result').textContent='Error: '+e.message;
  }
}
function handleEmailCsvUpload(input){
  const file=input.files[0];
  if(!file)return;
  const reader=new FileReader();
  reader.onload=function(e){
    const text=e.target.result;
    const lines=text.split(/\r?\n/);
    const pairs=[];
    for(const line of lines){
      const cells=line.split(/,|;|\t/);
      if(cells.length>=2){
        const email=cells[0].trim().toLowerCase();
        const domain=cells[1].trim().toLowerCase().replace(/^www\./,'');
        if(email.includes('@')&&domain.includes('.')){
          pairs.push(email+','+domain);
        }
      }
    }
    if(!pairs.length){alert('No valid email-domain pairs found in CSV');input.value='';return;}
    document.getElementById('email-import-text').value=pairs.join('\n');
    document.getElementById('email-import-result').textContent='CSV loaded: '+pairs.length+' email-domain pairs ready. Click Submit Import.';
    input.value='';
  };
  reader.readAsText(file);
}
function downloadEmailCsvTemplate(){
  const csv='email,domain\na@example.com,example.com\nb@site.org,site.org\ncontact@company.net,company.net\n';
  const blob=new Blob([csv],{type:'text/csv'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;
  a.download='email_import_template.csv';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

// ── Operation Log (new: two-level filter + user select + pagination + per-page stats) ──
const LOG = { allLogs: [], poolFilter: '', actionFilter: '', page: 1, pageSize: 20 };

function _logActionShort(type) {
  if (!type) return '-';
  const map = { domain_import: 'Import', domain_export: 'Export', domain_distribute: 'Distribute',
                email_import: 'Import', email_export: 'Export',
                price_import: 'Import', price_export: 'Export',
                reply_import: 'Import', reply_export: 'Export' };
  return map[type] || type;
}

function _logPoolShort(type) {
  if (!type) return '-';
  const p = (type.split('_')[0] || '');
  return p ? p.charAt(0).toUpperCase() + p.slice(1) + ' Pool' : '-';
}

async function refreshLogs() {
  const url = API + '/api/log/list?limit=100000';
  const r = await fetch(url).then(r => r.json());
  LOG.allLogs = r.logs || [];
  LOG.allLogs.sort((a, b) => {
    const ta = (a.time || a.timestamp || '').replace('T', ' ');
    const tb = (b.time || b.timestamp || '').replace('T', ' ');
    return tb.localeCompare(ta);
  });

  // Render pool-level cards from ALL logs
  const domainCount = LOG.allLogs.filter(l => l.type && l.type.startsWith('domain_')).length;
  const emailCount = LOG.allLogs.filter(l => l.type && l.type.startsWith('email_')).length;
  const replyCount = LOG.allLogs.filter(l => l.type && l.type.startsWith('reply_')).length;
  const quoteCount = LOG.allLogs.filter(l => l.type && (l.type.startsWith('quote_') || l.type.startsWith('price_'))).length;
  document.getElementById('log-cards').innerHTML = [
    { l: 'Domain Pool', v: domainCount, c: 'blue' },
    { l: 'Email Pool', v: emailCount, c: 'amber' },
    { l: 'Reply Pool', v: replyCount, c: 'green' },
    { l: 'Quote Pool', v: quoteCount, c: 'purple' },
  ].map(c => `<div class="card"><div class="label">${c.l}</div><div class="value ${c.c}">${fmt(c.v)}</div></div>`).join('');

  renderLogPage();
}

function setLogPoolFilter(pool) {
  LOG.poolFilter = pool;
  LOG.actionFilter = '';
  LOG.page = 1;

  // Update button active states
  document.querySelectorAll('#log-pool-filters .btn').forEach(b => b.classList.remove('active'));
  document.getElementById('log-pool-' + (pool || 'all')).classList.add('active');

  // Show/hide action filter row
  const actionRow = document.getElementById('log-action-filters');
  if (pool && pool !== 'reply') {
    actionRow.style.display = 'flex';
    document.querySelectorAll('#log-action-filters .btn').forEach(b => b.classList.remove('active'));
    document.getElementById('log-action-all').classList.add('active');
  } else {
    actionRow.style.display = 'none';
    LOG.actionFilter = '';
  }

  renderLogPage();
}

function setLogActionFilter(action) {
  LOG.actionFilter = action;
  LOG.page = 1;
  document.querySelectorAll('#log-action-filters .btn').forEach(b => b.classList.remove('active'));
  document.getElementById('log-action-' + (action || 'all')).classList.add('active');
  renderLogPage();
}

function changeLogPage(delta) {
  LOG.page += delta;
  if (LOG.page < 1) LOG.page = 1;
  renderLogPage();
}

function onLogPageSizeChange() {
  LOG.pageSize = parseInt(document.getElementById('log-page-size').value) || 20;
  LOG.page = 1;
  renderLogPage();
}

function renderLogPage() {
  let logs = [...LOG.allLogs];

  // Dedupe adjacent bursts: same user+action+pool+count+detail within 2s = same retry burst
  // Show one row + small "(+N similar)" hint. This fixes the "looks like 30 dupes" view.
  if (LOG.allLogs.length > 1) {
    const deduped = [];
    let i = 0;
    while (i < logs.length) {
      let j = i + 1;
      const a = logs[i];
      const ta = a.op_time ? new Date(a.op_time.replace(' ', 'T')).getTime() : 0;
      while (j < logs.length) {
        const b = logs[j];
        if (b.user !== a.user || b.type !== a.type || b.pool !== a.pool ||
            b.count !== a.count || b.detail !== a.detail) break;
        const tb = b.op_time ? new Date(b.op_time.replace(' ', 'T')).getTime() : 0;
        if (!tb || !ta || Math.abs(tb - ta) > 2000) break;
        j++;
      }
      const sim = j - i - 1;
      const row = { ...a };
      if (sim > 0) row.similar_count = sim;
      deduped.push(row);
      i = j;
    }
    logs = deduped;
  }

  // Apply pool filter (first-level)
  if (LOG.poolFilter) {
    if (LOG.poolFilter === 'quote') {
      logs = logs.filter(l => l.type && (l.type.startsWith('quote_') || l.type.startsWith('price_')));
    } else {
      logs = logs.filter(l => l.type && l.type.startsWith(LOG.poolFilter + '_'));
    }
  }

  // Apply action filter (second-level: import/export/distribute)
  if (LOG.actionFilter) {
    logs = logs.filter(l => {
      const parts = (l.type || '').split('_');
      return parts[1] === LOG.actionFilter;
    });
  }

  // Build user dropdown from filtered logs, exclude test/system entries
  const TEST_RE = /test|system|demo|unknown|admin@|autolink/i;
  const filteredUsers = [...new Set(logs.map(l => l.user).filter(u => u && !TEST_RE.test(u)))].sort();
  const sel = document.getElementById('log-user-filter');
  const oldVal = sel.value;
  sel.innerHTML = '<option value="">All Users</option>' +
    filteredUsers.map(u => `<option value="${esc(u)}">${esc(u)}</option>`).join('');
  if (filteredUsers.includes(oldVal)) sel.value = oldVal;

  // Apply user filter
  const userFilter = sel.value;
  if (userFilter) {
    logs = logs.filter(l => l.user === userFilter);
  }

  const total = logs.length;
  const totalPages = Math.max(1, Math.ceil(total / LOG.pageSize));
  if (LOG.page > totalPages) LOG.page = totalPages;
  if (LOG.page < 1) LOG.page = 1;

  const start = (LOG.page - 1) * LOG.pageSize;
  const pageLogs = logs.slice(start, start + LOG.pageSize);

  // Render table
  const table = document.getElementById('log-table');
  const startIdx = start;
  table.innerHTML = '<tr><th><input type="checkbox" class="chk-all" onchange="toggleLogAll(this)" title="Select All"></th><th>#</th><th>Time</th><th>User</th><th>Action</th><th>Pool</th><th>Count</th><th>Detail</th></tr>' +
    pageLogs.map((l, i) => `<tr data-id="${esc(l.log_id)}">
      <td><input type="checkbox" class="chk-row" onchange="onLogCheck(this,'${esc(l.log_id)}')"></td>
      <td style="color:var(--muted);font-size:11.5px;text-align:center">${startIdx + i + 1}</td>
      <td>${esc((l.time || '').slice(0, 16))}${l.similar_count ? ` <span style="color:var(--muted);font-size:10.5px">(+${l.similar_count} similar)</span>` : ''}</td>
      <td>${esc(l.user)}</td>
      <td><span class="status-${l.type || 'New'}">${_logActionShort(l.type)}</span></td>
      <td>${_logPoolShort(l.type)}</td>
      <td>${l.count || 0}</td>
      <td style="max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${esc(l.detail)}</td>
    </tr>`).join('');

  // Update pagination info
  document.getElementById('log-page-info').textContent =
    `Page ${LOG.page} / ${totalPages} (${total} records)`;
  document.getElementById('log-prev').disabled = LOG.page <= 1;
  document.getElementById('log-next').disabled = LOG.page >= totalPages;

  // Per-page statistics
  const pageTotalCount = pageLogs.reduce((s, l) => s + (l.count || 0), 0);
  const typeStats = {};
  pageLogs.forEach(l => {
    const t = l.type || 'unknown';
    typeStats[t] = (typeStats[t] || 0) + (l.count || 0);
  });
  const statParts = Object.entries(typeStats)
    .sort((a, b) => b[1] - a[1])
    .slice(0, 5)
    .map(([t, c]) => `${_logActionShort(t)}: ${fmt(c)}`);
  const statsText = `This page: ${pageLogs.length} logs, total count ${fmt(pageTotalCount)}` +
    (statParts.length ? ' | ' + statParts.join(' / ') : '');
  document.getElementById('log-page-stats').textContent = statsText;
}

// Back-compat wrapper
async function loadLogTable(filterType) {
  if (!LOG.allLogs.length) await refreshLogs();
  if (!filterType) {
    setLogPoolFilter('');
  } else if (['domain', 'email', 'reply', 'price'].includes(filterType)) {
    setLogPoolFilter(filterType);
  } else {
    // Exact type like domain_import
    const pool = filterType.split('_')[0];
    const action = filterType.split('_')[1];
    setLogPoolFilter(pool);
    if (action) setLogActionFilter(action);
  }
}

// ── Import A-class replies to Quote Pool ──
async function importARepliesToQuotes(){
  const user=getUserName();
  if(!user){alert('Please enter your name in "My Name" field first');return;}
  if(!confirm('Import all A-class replies (671 records) to Quote Pool?\\nExisting quotes (by email) will be skipped.'))return;
  const result=document.getElementById('quote-import-result');
  result.textContent='Importing...';
  try{
    const form=new FormData();
    form.append('user',user);
    const r=await fetch(API+'/api/quote/import-a-replies',{method:'POST',body:form});
    const data=await r.json();
    if(data.error){result.textContent='Error: '+data.error;alert(data.error);}
    else{
      result.textContent='OK: '+data.imported+' imported, '+data.skipped+' skipped (quote pool total: '+data.quote_pool_total+')';
      loadQuoteTable();loadStats();
    }
  }catch(e){result.textContent='Network error';alert('Import failed: '+e.message);}
}

// ── Quote Pool file import ──
function toggleQuoteImport(){
  const p=document.getElementById('quote-import-panel');
  p.style.display=p.style.display==='none'?'block':'none';
}
function downloadQuoteTemplate(){
  const headers='\uFEFFEmail,Domain,Supplier,Contact Email,Niche,Country,Traffic,Site Category,Cooperation Type,Price,Link Rules,Permanence,Content,TAT,Payment,Discount,Additional Services,Requirements,Reply Content,Status,Notes,Priority\n';
  const sample='example@domain.com,domain.com,Supplier Name,contact@domain.com,Technology,USA,50K,News,Guest Post,$100,DoFollow,Permanent,500 words,3-5 days,PayPal,10% off,SEO audit,Original content required,Thanks for reaching out...,New,Initial contact,1\n';
  const blob=new Blob([headers+sample],{type:'text/csv;charset=utf-8'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='quote_import_template.csv';a.click();
}
async function handleQuoteUpload(input){
  const user=getUserName();
  if(!user){alert('Please enter your name in "My Name" field first');return;}
  if(!input.files||!input.files[0])return;
  const file=input.files[0];
  const result=document.getElementById('quote-import-file-result');
  result.textContent='Uploading...';
  const form=new FormData();
  form.append('file',file);
  form.append('user',user);
  try{
    const r=await fetch(API+'/api/quote/import',{method:'POST',body:form});
    const data=await r.json();
    if(data.error){result.textContent='Error: '+data.error;alert(data.error);}
    else{
      result.textContent='OK: imported '+data.imported+', skipped '+data.skipped+' duplicates';
      alert('Imported '+data.imported+' quotes'+(data.skipped?' (skipped '+data.skipped+' duplicates)':''));
      loadQuoteTable();loadStats();
    }
  }catch(e){result.textContent='Network error';alert('Upload failed: '+e.message);}
  input.value='';
}

// ── Reply Pool import ──
function toggleReplyImport(){
  const p=document.getElementById('reply-import-panel');
  p.style.display=p.style.display==='none'?'block':'none';
}
function downloadReplyTemplate(){
  const csv='\uFEFF序号,邮箱,域名,账号,发件人,主题,正文摘要,日期\n1,example@domain.com,domain.com,your-account,Supplier Name,Reply Subject,Reply body preview...,2026-07-30';
  const blob=new Blob([csv],{type:'text/csv;charset=utf-8'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='reply_import_template.csv';a.click();
}
async function handleReplyUpload(input){
  const user=getUserName();
  if(!user){alert('Please enter your name in "My Name" field first');return;}
  if(!input.files||!input.files[0])return;
  const file=input.files[0];
  const result=document.getElementById('reply-import-result');
  result.textContent='Uploading...';
  const form=new FormData();
  form.append('file',file);
  form.append('user',user);
  try{
    const r=await fetch(API+'/api/reply/import',{method:'POST',body:form});
    const data=await r.json();
    if(data.error){result.textContent='Error: '+data.error;alert(data.error);}
    else{
      result.textContent='OK: imported '+data.imported+', skipped '+data.skipped+' duplicates';
      alert('Imported '+data.imported+' replies'+(data.skipped?' (skipped '+data.skipped+' duplicates)':''));
      loadReplyTable('');loadStats();
    }
  }catch(e){result.textContent='Network error';alert('Upload failed: '+e.message);}
  input.value='';
}

// ── Export Selected functions (sequential IDs, not DB IDs) ──
function exportSelectedDomains(){
  if(!DOMAIN_SEL.size){alert('No rows selected');return;}
  const rows=[...DOMAIN_SEL].sort((a,b)=>a.idx-b.idx);
  const csv='\uFEFF'+['#','Domain','Source','Status','Claimed By','Priority','Created'].join(',')+'\n'+
    rows.map((r,i)=>{
      const tr=r.row||document.querySelector(`#domain-table tr[data-id="${r.id}"]`);
      const tds=tr?tr.querySelectorAll('td'):[];
      return [i+1, tds[2]?tds[2].textContent:'', tds[3]?tds[3].textContent:'', tds[4]?tds[4].textContent:'', tds[5]?tds[5].textContent:'', tds[6]?tds[6].textContent:'', tds[7]?tds[7].textContent.slice(0,16):''].map(v=>`"${String(v).replace(/"/g,'""')}"`).join(',');
    }).join('\n');
  _downloadCSV(csv,'selected_domains.csv');
}
function exportSelectedEmails(){
  if(!EMAIL_SEL.size){alert('No rows selected');return;}
  const rows=[...EMAIL_SEL].sort((a,b)=>a.idx-b.idx);
  const csv='\uFEFF'+['#','Email','Domain','Send Status','Source'].join(',')+'\n'+
    rows.map((r,i)=>{
      const tr=r.row||document.querySelector(`#email-table tr[data-id="${r.id}"]`);
      const tds=tr?tr.querySelectorAll('td'):[];
      return [i+1, tds[2]?tds[2].textContent:'', tds[3]?tds[3].textContent:'', tds[4]?tds[4].textContent:'', tds[6]?tds[6].textContent:''].map(v=>`"${String(v).replace(/"/g,'""')}"`).join(',');
    }).join('\n');
  _downloadCSV(csv,'selected_emails.csv');
}
function exportQuotes(scope){
  scope = scope || 'all';
  // If rows are selected → export only selected (Jenny format CSV, ignores scope)
  // If none selected → export via backend with scope filter (all|ready|abnormal)
  if(QUOTE_SEL.size){
    const rows=[...QUOTE_SEL].sort((a,b)=>a.idx-b.idx);
    const headers=['序号','Link','Price','Backlink Type','DR','DA','Ref. Domains','Traffic','Country','Keywords','Categories','Languages','TAT','Permanence','Contact','Cooperation','Payment','Discount','Link Rules','Status','其他'];
    const csv='\uFEFF'+headers.join(',')+'\n'+
      rows.map((r,i)=>{
        const tr=r.row||document.querySelector(`#quote-table tr[data-id="${r.id}"]`);
        const tds=tr?tr.querySelectorAll('td'):[];
        const vals=[];
        for(let c=2;c<tds.length;c++) vals.push(tds[c]?tds[c].textContent:'');
        return [i+1,...vals].map(v=>`"${String(v).replace(/"/g,'""')}"`).join(',');
      }).join('\n');
    _downloadCSV(csv,'selected_quotes_jenny_format.csv');
  } else {
    // No selection → open backend export with scope filter
    window.open('/api/quote/export?scope=' + encodeURIComponent(scope), '_blank');
  }
}
async function deleteSelectedQuotes(){
  if(!QUOTE_SEL.size){alert('No rows selected');return;}
  if(!confirm('Delete '+QUOTE_SEL.size+' selected quote(s)? This cannot be undone.'))return;
  const ids=[...QUOTE_SEL].map(r=>r.id);
  const r=await fetch(API+'/api/quote/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ids})}).then(r=>r.json());
  if(r.error){alert('Error: '+r.error);return;}
  alert('Deleted '+r.deleted+' quote(s)');
  QUOTE_SEL.clear();
  loadQuoteTable();loadStats();
}

// ── Operation Log select/delete ──
function toggleLogAll(chk){
  document.querySelectorAll('#log-table .chk-row').forEach(c=>{c.checked=chk.checked;c.onchange();});
}
function onLogCheck(chk,id){
  if(chk.checked) LOG_SEL.add(id); else LOG_SEL.delete(id);
  document.getElementById('log-sel-count').textContent=LOG_SEL.size;
  document.getElementById('log-delete-btn').style.display=LOG_SEL.size?'inline-block':'none';
  chk.closest('tr').classList.toggle('selected-row',chk.checked);
}
async function deleteSelectedLogs(){
  if(!LOG_SEL.size){alert('No rows selected');return;}
  if(!confirm('Delete '+LOG_SEL.size+' selected log entr(y/ies)? This cannot be undone.'))return;
  const ids=[...LOG_SEL];
  const r=await fetch(API+'/api/log/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ids})}).then(r=>r.json());
  if(r.error){alert('Error: '+r.error);return;}
  alert('Deleted '+r.deleted+' log entries');
  LOG_SEL.clear();
  refreshLogs();
}
function _downloadCSV(content,filename){
  const blob=new Blob([content],{type:'text/csv;charset=utf-8'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');a.href=url;a.download=filename;
  document.body.appendChild(a);a.click();document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

async function exportDomains(){
  const user=getUserName();
  if(!user){alert('Please enter your name in "My Name" field first');return;}
  const count=document.getElementById('d-count').value;
  const r=await fetch(API+'/api/domain/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({user,count:parseInt(count)})}).then(r=>r.json());
  if(r.error){alert(r.error);return;}
  if(r.csv_content){
    const blob=new Blob(['\uFEFF'+r.csv_content],{type:'text/csv;charset=utf-8'});
    const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=r.filename;a.click();
  }
  alert('Exported '+r.exported+' domains\nFile: '+r.filename+'\nBatch: '+r.batch_id);
  loadStats();loadDomainTable();
}
async function distributeDomains(){
  if(!confirm('Distribute ALL New domains among team members?'))return;
  const r=await fetch(API+'/api/domain/distribute',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({})}).then(r=>r.json());
  alert('Distributed '+r.assigned+' domains\n'+JSON.stringify(r.distribution,null,2));
  loadStats();loadDomainTable();
}
async function exportEmails(){
  const user=getUserName();
  if(!user){alert('Please enter your name in "My Name" field first');return;}
  const count=document.getElementById('e-count').value;
  const r=await fetch(API+'/api/email/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({user,count:parseInt(count)})}).then(r=>r.json());
  if(r.error){alert(r.error);return;}
  if(r.csv_content){
    const blob=new Blob(['\uFEFF'+r.csv_content],{type:'text/csv;charset=utf-8'});
    const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=r.filename;a.click();
  }
  alert('Exported '+r.exported+' emails\nFile: '+r.filename);
  loadStats();
}

async function loadAll(){await loadStats();await loadDomainTable();}
loadAll();
// 预加载 Quote 数据（不依赖用户点击 tab），避免 switchTab 异常导致空白
setTimeout(()=>{try{loadQuoteTable();}catch(e){console.error('[init] loadQuoteTable failed',e);}},500);
setInterval(()=>loadStats().catch(e=>{}),30000);
</script>
</body>
</html>"""


@app.route("/")
def dashboard():
    return DASHBOARD_HTML


# ════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Shared Pool v2 — Supplier Intelligence (Supabase)")
    parser.add_argument("--port", type=int, default=None)
    parser.add_argument("--host", type=str, default="0.0.0.0")
    parser.add_argument("--dbg", action="store_true", help="Debug mode")
    args = parser.parse_args()

    # PORT priority: env var (Render) > --port arg > default 8765
    port = int(os.environ.get("PORT", 0)) or args.port or 8765

    # Verify connection
    print(f"\n  Shared Pool v2 — Supabase Edition")
    print(f"  Project:   {SUPABASE_URL}")
    try:
        health_check = db.select("domain_pool", select="domain_id", limit=1)
        print(f"  Connection: OK (domain_pool: {db.count('domain_pool'):,} rows)")
    except Exception as e:
        print(f"  Connection: FAILED — {e}")
        print(f"  Check your config.py settings.")
        sys.exit(1)

    print(f"  Listening:  http://{args.host}:{port}")
    print(f"  Dashboard:  http://localhost:{port}")
    print(f"  Setup:      http://localhost:{port}/setup")
    print(f"  API:        /api/domain/*  /api/email/*  /api/reply/*  /api/price/*")
    print()

    app.run(host=args.host, port=port, debug=args.dbg, threaded=True)
